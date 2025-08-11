#!/usr/bin/env python3
"""
🤖 Ultron Analyzer - Core Analysis Engine
Main analyzer class for Ultron site performance analysis
"""

import requests
import time
import json
import re
import sys
from urllib.parse import urljoin, urlparse, parse_qs
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, asdict
from typing import List, Dict, Tuple, Optional
import warnings
from datetime import datetime
import os

# Excel generation imports
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import PieChart, BarChart, Reference
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# Suppress SSL warnings for testing
warnings.filterwarnings('ignore', message='Unverified HTTPS request')

@dataclass
class PerformanceMetrics:
    """Container for performance metrics"""
    dns_time: float = 0.0
    connect_time: float = 0.0
    ssl_time: float = 0.0
    response_time: float = 0.0
    download_time: float = 0.0
    total_time: float = 0.0
    page_size: int = 0
    status_code: int = 0

@dataclass
class ImageInfo:
    """Container for image information"""
    url: str
    size_bytes: int = 0
    dimensions: Optional[Tuple[int, int]] = None
    format: str = ""
    is_optimized: bool = False
    issues: List[str] = None

    def __post_init__(self):
        if self.issues is None:
            self.issues = []

@dataclass
class LinkInfo:
    """Container for link information"""
    url: str
    status_code: int
    response_time: float
    is_broken: bool
    error_message: str = ""

@dataclass
class SEOMetrics:
    """Container for SEO metrics"""
    title: str = ""
    meta_description: str = ""
    h1_tags: List[str] = None
    h2_tags: List[str] = None
    images_without_alt: int = 0
    internal_links: int = 0
    external_links: int = 0

    def __post_init__(self):
        if self.h1_tags is None:
            self.h1_tags = []
        if self.h2_tags is None:
            self.h2_tags = []

class UltronAnalyzer:
    """🤖 Ultron - Advanced site performance and quality analyzer"""
    
    def __init__(self, timeout: int = 30, max_workers: int = 10):
        self.timeout = timeout
        self.max_workers = max_workers
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        })
        
    def get_page_performance(self, url: str) -> PerformanceMetrics:
        """Measure detailed page performance metrics"""
        try:
            start_time = time.time()
            response = self.session.get(url, timeout=self.timeout, verify=False)
            end_time = time.time()
            
            metrics = PerformanceMetrics(
                total_time=end_time - start_time,
                response_time=end_time - start_time,
                page_size=len(response.content),
                status_code=response.status_code
            )
            
            # Try to get more detailed timing if available
            if hasattr(response, 'elapsed'):
                metrics.response_time = response.elapsed.total_seconds()
            
            return metrics
            
        except requests.RequestException as e:
            print(f"Error measuring performance for {url}: {e}")
            return PerformanceMetrics()
    
    def check_security_headers(self, url: str) -> Dict[str, bool]:
        """Check for important security headers"""
        security_headers = {
            'X-Content-Type-Options': False,
            'X-Frame-Options': False,
            'X-XSS-Protection': False,
            'Strict-Transport-Security': False,
            'Content-Security-Policy': False,
            'Referrer-Policy': False,
            'Permissions-Policy': False
        }
        
        try:
            response = self.session.head(url, timeout=self.timeout, verify=False)
            headers = response.headers
            
            for header in security_headers:
                if header in headers:
                    security_headers[header] = True
                    
        except requests.RequestException as e:
            print(f"Error checking security headers for {url}: {e}")
            
        return security_headers
    
    def extract_links(self, url: str, html_content: str) -> List[str]:
        """Extract all links from HTML content"""
        try:
            from bs4 import BeautifulSoup
            soup = BeautifulSoup(html_content, 'html.parser')
            links = []
            
            for link in soup.find_all(['a', 'link'], href=True):
                href = link.get('href')
                if href:
                    absolute_url = urljoin(url, href)
                    if absolute_url.startswith(('http://', 'https://')):
                        links.append(absolute_url)
                        
            return list(set(links))  # Remove duplicates
            
        except ImportError:
            print("BeautifulSoup not available. Install with: pip install beautifulsoup4")
            return []
        except Exception as e:
            print(f"Error extracting links: {e}")
            return []
    
    def extract_images(self, url: str, html_content: str) -> List[str]:
        """Extract all image URLs from HTML content"""
        try:
            from bs4 import BeautifulSoup
            soup = BeautifulSoup(html_content, 'html.parser')
            images = []
            
            for img in soup.find_all('img', src=True):
                src = img.get('src')
                if src:
                    absolute_url = urljoin(url, src)
                    if absolute_url.startswith(('http://', 'https://')):
                        images.append(absolute_url)
                        
            return list(set(images))  # Remove duplicates
            
        except ImportError:
            print("BeautifulSoup not available. Install with: pip install beautifulsoup4")
            return []
        except Exception as e:
            print(f"Error extracting images: {e}")
            return []
    
    def check_link(self, url: str) -> LinkInfo:
        """Check if a single link is working"""
        start_time = time.time()
        try:
            response = self.session.head(url, timeout=self.timeout, verify=False, allow_redirects=True)
            response_time = time.time() - start_time
            
            is_broken = response.status_code >= 400
            return LinkInfo(
                url=url,
                status_code=response.status_code,
                response_time=response_time,
                is_broken=is_broken
            )
            
        except requests.RequestException as e:
            response_time = time.time() - start_time
            return LinkInfo(
                url=url,
                status_code=0,
                response_time=response_time,
                is_broken=True,
                error_message=str(e)
            )
    
    def check_broken_links(self, links: List[str]) -> List[LinkInfo]:
        """Check multiple links for broken status using threading"""
        link_results = []
        
        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            future_to_url = {executor.submit(self.check_link, url): url for url in links[:50]}  # Limit to 50 links
            
            for future in as_completed(future_to_url):
                result = future.result()
                link_results.append(result)
                
        return link_results
    
    def analyze_image(self, img_url: str) -> ImageInfo:
        """Analyze a single image for optimization"""
        try:
            response = self.session.head(img_url, timeout=self.timeout, verify=False)
            size_bytes = int(response.headers.get('content-length', 0))
            content_type = response.headers.get('content-type', '')
            
            image_info = ImageInfo(
                url=img_url,
                size_bytes=size_bytes,
                format=content_type
            )
            
            # Check for optimization issues
            if size_bytes > 500000:  # 500KB
                image_info.issues.append(f"Large file size: {size_bytes / 1024:.1f}KB")
            
            if 'webp' not in content_type.lower() and 'svg' not in content_type.lower():
                image_info.issues.append("Not using modern format (WebP/SVG)")
            
            image_info.is_optimized = len(image_info.issues) == 0
            
            return image_info
            
        except requests.RequestException as e:
            return ImageInfo(
                url=img_url,
                issues=[f"Error loading image: {e}"]
            )
    
    def analyze_images(self, image_urls: List[str]) -> List[ImageInfo]:
        """Analyze multiple images using threading"""
        image_results = []
        
        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            future_to_url = {executor.submit(self.analyze_image, url): url for url in image_urls[:20]}  # Limit to 20 images
            
            for future in as_completed(future_to_url):
                result = future.result()
                image_results.append(result)
                
        return image_results
    
    def analyze_seo(self, url: str, html_content: str) -> SEOMetrics:
        """Analyze basic SEO metrics"""
        try:
            from bs4 import BeautifulSoup
            soup = BeautifulSoup(html_content, 'html.parser')
            
            seo_metrics = SEOMetrics()
            
            # Title
            title_tag = soup.find('title')
            if title_tag:
                seo_metrics.title = title_tag.get_text().strip()
            
            # Meta description
            meta_desc = soup.find('meta', attrs={'name': 'description'})
            if meta_desc:
                seo_metrics.meta_description = meta_desc.get('content', '').strip()
            
            # Heading tags
            seo_metrics.h1_tags = [h1.get_text().strip() for h1 in soup.find_all('h1')]
            seo_metrics.h2_tags = [h2.get_text().strip() for h2 in soup.find_all('h2')]
            
            # Images without alt text
            images = soup.find_all('img')
            seo_metrics.images_without_alt = len([img for img in images if not img.get('alt')])
            
            # Count links
            domain = urlparse(url).netloc
            for link in soup.find_all('a', href=True):
                href = link.get('href')
                if href:
                    if href.startswith(('http://', 'https://')):
                        if domain in href:
                            seo_metrics.internal_links += 1
                        else:
                            seo_metrics.external_links += 1
                    else:
                        seo_metrics.internal_links += 1
            
            return seo_metrics
            
        except ImportError:
            print("BeautifulSoup not available for SEO analysis")
            return SEOMetrics()
        except Exception as e:
            print(f"Error in SEO analysis: {e}")
            return SEOMetrics()
    
    def check_mobile_friendly(self, html_content: str) -> Dict[str, bool]:
        """Check basic mobile-friendly indicators"""
        mobile_checks = {
            'viewport_meta': False,
            'responsive_images': False,
            'touch_friendly': False
        }
        
        try:
            from bs4 import BeautifulSoup
            soup = BeautifulSoup(html_content, 'html.parser')
            
            # Viewport meta tag
            viewport = soup.find('meta', attrs={'name': 'viewport'})
            if viewport:
                mobile_checks['viewport_meta'] = True
            
            # Responsive images
            images = soup.find_all('img')
            responsive_images = [img for img in images if img.get('srcset') or 'responsive' in img.get('class', [])]
            if responsive_images:
                mobile_checks['responsive_images'] = True
            
            # Touch-friendly elements (basic check)
            buttons = soup.find_all(['button', 'a'])
            if len(buttons) > 0:
                mobile_checks['touch_friendly'] = True
                
        except ImportError:
            pass
        except Exception as e:
            print(f"Error in mobile-friendly check: {e}")
            
        return mobile_checks
    
    def generate_insights(self, performance: PerformanceMetrics, security: Dict[str, bool], 
                         seo: SEOMetrics, images: List[ImageInfo], links: List[LinkInfo],
                         mobile: Dict[str, bool]) -> List[str]:
        """Generate actionable insights based on analysis"""
        insights = []
        
        # Performance insights
        if performance.total_time > 3.0:
            insights.append(f"⚠️ CRITICAL: Page load time is {performance.total_time:.2f}s (should be <3s)")
        elif performance.total_time > 1.5:
            insights.append(f"⚡ Page load time is {performance.total_time:.2f}s (aim for <1.5s for optimal UX)")
        
        if performance.page_size > 2000000:  # 2MB
            insights.append(f"📦 Large page size: {performance.page_size / 1024 / 1024:.1f}MB (consider optimizing)")
        
        # Security insights
        missing_headers = [header for header, present in security.items() if not present]
        if missing_headers:
            insights.append(f"🔒 Missing security headers: {', '.join(missing_headers)}")
        
        # SEO insights
        if not seo.title:
            insights.append("📝 Missing page title")
        elif len(seo.title) > 60:
            insights.append(f"📝 Title too long: {len(seo.title)} chars (aim for <60)")
        
        if not seo.meta_description:
            insights.append("📄 Missing meta description")
        elif len(seo.meta_description) > 155:
            insights.append(f"📄 Meta description too long: {len(seo.meta_description)} chars (aim for <155)")
        
        if len(seo.h1_tags) == 0:
            insights.append("🏷️ No H1 tags found")
        elif len(seo.h1_tags) > 1:
            insights.append(f"🏷️ Multiple H1 tags found ({len(seo.h1_tags)}) - use only one H1 per page")
        
        if seo.images_without_alt > 0:
            insights.append(f"♿ {seo.images_without_alt} images missing alt text (accessibility issue)")
        
        # Image insights
        large_images = [img for img in images if img.size_bytes > 500000]
        if large_images:
            insights.append(f"🖼️ {len(large_images)} large images found (>500KB) - consider compression")
        
        unoptimized_images = [img for img in images if not img.is_optimized]
        if unoptimized_images:
            insights.append(f"🔧 {len(unoptimized_images)} images could be optimized")
        
        # Link insights
        broken_links = [link for link in links if link.is_broken]
        if broken_links:
            insights.append(f"🔗 {len(broken_links)} broken links found - fix immediately")
        
        # Mobile insights
        if not mobile['viewport_meta']:
            insights.append("📱 Missing viewport meta tag - page may not be mobile-friendly")
        
        if not mobile['responsive_images']:
            insights.append("📱 No responsive images detected - consider using srcset")
        
        return insights
    
    def run_comprehensive_check(self, url: str) -> Dict:
        """Run all checks and return comprehensive results"""
        print(f"🔍 Starting comprehensive analysis of: {url}")
        print("=" * 60)
        
        results = {
            'url': url,
            'timestamp': datetime.now().isoformat(),
            'performance': {},
            'security': {},
            'seo': {},
            'images': [],
            'links': [],
            'mobile': {},
            'insights': []
        }
        
        try:
            # Get page content
            print("📄 Fetching page content...")
            response = self.session.get(url, timeout=self.timeout, verify=False)
            html_content = response.text
            
            # Performance analysis
            print("⚡ Analyzing performance...")
            performance = self.get_page_performance(url)
            results['performance'] = asdict(performance)
            
            # Security headers
            print("🔒 Checking security headers...")
            security = self.check_security_headers(url)
            results['security'] = security
            
            # SEO analysis
            print("📝 Analyzing SEO...")
            seo = self.analyze_seo(url, html_content)
            results['seo'] = asdict(seo)
            
            # Extract and check links
            print("🔗 Extracting and checking links...")
            links = self.extract_links(url, html_content)
            print(f"   Found {len(links)} links, checking status...")
            link_results = self.check_broken_links(links)
            results['links'] = [asdict(link) for link in link_results]
            
            # Extract and analyze images
            print("🖼️ Analyzing images...")
            images = self.extract_images(url, html_content)
            print(f"   Found {len(images)} images, analyzing...")
            image_results = self.analyze_images(images)
            results['images'] = [asdict(img) for img in image_results]
            
            # Mobile-friendly check
            print("📱 Checking mobile-friendliness...")
            mobile = self.check_mobile_friendly(html_content)
            results['mobile'] = mobile
            
            # Generate insights
            print("💡 Generating insights...")
            insights = self.generate_insights(performance, security, seo, image_results, link_results, mobile)
            results['insights'] = insights
            
            return results
            
        except requests.RequestException as e:
            print(f"❌ Error accessing {url}: {e}")
            results['error'] = str(e)
            return results
    
    def print_results(self, results: Dict):
        """Print formatted results to console"""
        print("\n" + "🤖" * 20)
        print("📊 ULTRON ANALYSIS RESULTS")
        print("🤖" * 20)
        
        if 'error' in results:
            print(f"❌ Error: {results['error']}")
            return
        
        # Performance Summary
        perf = results['performance']
        print(f"\n⚡ PERFORMANCE SUMMARY")
        print(f"   Load Time: {perf['total_time']:.2f}s")
        print(f"   Page Size: {perf['page_size'] / 1024:.1f}KB")
        print(f"   Status Code: {perf['status_code']}")
        
        # Security Summary
        security = results['security']
        secure_headers = sum(security.values())
        total_headers = len(security)
        print(f"\n🔒 SECURITY SUMMARY")
        print(f"   Security Headers: {secure_headers}/{total_headers}")
        
        # SEO Summary
        seo = results['seo']
        print(f"\n📝 SEO SUMMARY")
        print(f"   Title: {'✅' if seo['title'] else '❌'} {seo['title'][:50]}...")
        print(f"   Meta Description: {'✅' if seo['meta_description'] else '❌'}")
        print(f"   H1 Tags: {len(seo['h1_tags'])}")
        print(f"   Images without Alt: {seo['images_without_alt']}")
        
        # Links Summary
        links = results['links']
        broken_links = [link for link in links if link['is_broken']]
        print(f"\n🔗 LINKS SUMMARY")
        print(f"   Total Links Checked: {len(links)}")
        print(f"   Broken Links: {len(broken_links)}")
        
        # Images Summary
        images = results['images']
        unoptimized = [img for img in images if not img['is_optimized']]
        print(f"\n🖼️ IMAGES SUMMARY")
        print(f"   Total Images: {len(images)}")
        print(f"   Need Optimization: {len(unoptimized)}")
        
        # Mobile Summary
        mobile = results['mobile']
        mobile_score = sum(mobile.values())
        print(f"\n📱 MOBILE SUMMARY")
        print(f"   Mobile Score: {mobile_score}/{len(mobile)}")
        
        # Insights
        insights = results['insights']
        print(f"\n💡 ACTIONABLE INSIGHTS ({len(insights)} issues found)")
        for i, insight in enumerate(insights, 1):
            print(f"   {i}. {insight}")
        
        if not insights:
            print("   🎉 Great! No major issues found!")
    
    def save_report(self, results: Dict, format: str = 'json'):
        """Save results to file"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        domain = urlparse(results['url']).netloc
        
        if format == 'json':
            filename = f"ultron_report_{domain}_{timestamp}.json"
            with open(filename, 'w') as f:
                json.dump(results, f, indent=2)
            print(f"\n💾 Report saved as: {filename}")
        
        elif format == 'html':
            filename = f"ultron_report_{domain}_{timestamp}.html"
            html_report = self.generate_html_report(results)
            with open(filename, 'w') as f:
                f.write(html_report)
            print(f"\n💾 HTML report saved as: {filename}")
            
        elif format == 'excel':
            if not EXCEL_AVAILABLE:
                print("❌ Excel generation not available. Install openpyxl: pip install openpyxl")
                return
            filename = f"ultron_report_{domain}_{timestamp}.xlsx"
            self.generate_excel_report(results, filename)
            print(f"\n💾 Excel report saved as: {filename}")
    
    def generate_excel_report(self, results: Dict, filename: str):
        """Generate comprehensive Excel report with multiple sheets"""
        if not EXCEL_AVAILABLE:
            return
        
        wb = Workbook()
        
        # Define styles
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid")
        subheader_font = Font(bold=True, size=12)
        subheader_fill = PatternFill(start_color="AED6F1", end_color="AED6F1", fill_type="solid")
        
        critical_fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
        warning_fill = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
        good_fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
        
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Sheet 1: Executive Summary
        ws_summary = wb.active
        ws_summary.title = "Executive Summary"
        
        # Header
        ws_summary.merge_cells('A1:E1')
        ws_summary['A1'] = f"🤖 Ultron Performance Report - {results['url']}"
        ws_summary['A1'].font = header_font
        ws_summary['A1'].fill = header_fill
        ws_summary['A1'].alignment = Alignment(horizontal='center')
        
        ws_summary['A2'] = f"Generated: {results['timestamp']}"
        ws_summary['A2'].font = Font(italic=True)
        
        # Overall Scores
        row = 4
        ws_summary[f'A{row}'] = "OVERALL PERFORMANCE SUMMARY"
        ws_summary[f'A{row}'].font = subheader_font
        ws_summary[f'A{row}'].fill = subheader_fill
        
        row += 2
        performance = results['performance']
        
        # Performance metrics
        metrics = [
            ("Load Time", f"{performance['total_time']:.2f}s", self._get_performance_color(performance['total_time'], 'time')),
            ("Page Size", f"{performance['page_size'] / 1024:.1f}KB", self._get_performance_color(performance['page_size'], 'size')),
            ("Status Code", performance['status_code'], good_fill if performance['status_code'] == 200 else critical_fill),
            ("Security Headers", f"{sum(results['security'].values())}/7", self._get_performance_color(sum(results['security'].values()), 'security')),
            ("Broken Links", len([l for l in results['links'] if l['is_broken']]), critical_fill if any(l['is_broken'] for l in results['links']) else good_fill),
            ("Images Need Optimization", len([i for i in results['images'] if not i['is_optimized']]), warning_fill if any(not i['is_optimized'] for i in results['images']) else good_fill),
            ("Mobile Score", f"{sum(results['mobile'].values())}/3", self._get_performance_color(sum(results['mobile'].values()), 'mobile')),
            ("Total Issues Found", len(results['insights']), critical_fill if len(results['insights']) > 5 else warning_fill if len(results['insights']) > 0 else good_fill)
        ]
        
        for metric, value, fill in metrics:
            ws_summary[f'A{row}'] = metric
            ws_summary[f'B{row}'] = value
            ws_summary[f'B{row}'].fill = fill
            ws_summary[f'A{row}'].border = border
            ws_summary[f'B{row}'].border = border
            row += 1
        
        # Sheet 2: Performance Details
        ws_perf = wb.create_sheet("Performance Analysis")
        self._create_performance_sheet(ws_perf, results, header_font, header_fill, subheader_font, subheader_fill, border)
        
        # Sheet 3: Security Analysis
        ws_security = wb.create_sheet("Security Analysis")
        self._create_security_sheet(ws_security, results, header_font, header_fill, subheader_font, subheader_fill, border)
        
        # Sheet 4: SEO Analysis
        ws_seo = wb.create_sheet("SEO Analysis")
        self._create_seo_sheet(ws_seo, results, header_font, header_fill, subheader_font, subheader_fill, border)
        
        # Sheet 5: Links Analysis
        ws_links = wb.create_sheet("Links Analysis")
        self._create_links_sheet(ws_links, results, header_font, header_fill, subheader_font, subheader_fill, border)
        
        # Sheet 6: Images Analysis
        ws_images = wb.create_sheet("Images Analysis")
        self._create_images_sheet(ws_images, results, header_font, header_fill, subheader_font, subheader_fill, border)
        
        # Sheet 7: Action Items
        ws_actions = wb.create_sheet("Action Items")
        self._create_action_items_sheet(ws_actions, results, header_font, header_fill, subheader_font, subheader_fill, border)
        
        # Auto-adjust column widths
        for ws in wb.worksheets:
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(filename)
    
    def _get_performance_color(self, value, metric_type):
        """Get color based on performance value"""
        if metric_type == 'time':
            if value <= 1.5:
                return PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")  # Green
            elif value <= 3.0:
                return PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")  # Orange
            else:
                return PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")  # Red
        elif metric_type == 'size':
            if value <= 1024000:  # 1MB
                return PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
            elif value <= 2048000:  # 2MB
                return PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
            else:
                return PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
        elif metric_type == 'security':
            if value >= 6:
                return PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
            elif value >= 4:
                return PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
            else:
                return PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
        elif metric_type == 'mobile':
            if value == 3:
                return PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
            elif value == 2:
                return PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
            else:
                return PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
    
    def _create_performance_sheet(self, ws, results, header_font, header_fill, subheader_font, subheader_fill, border):
        """Create performance analysis sheet"""
        ws['A1'] = "Performance Analysis"
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        
        perf = results['performance']
        
        row = 3
        performance_data = [
            ("Metric", "Value", "Benchmark", "Status"),
            ("DNS Lookup Time", f"{perf.get('dns_time', 0):.3f}s", "< 0.1s", ""),
            ("Connection Time", f"{perf.get('connect_time', 0):.3f}s", "< 0.1s", ""),
            ("SSL Handshake", f"{perf.get('ssl_time', 0):.3f}s", "< 0.1s", ""),
            ("Response Time", f"{perf['total_time']:.3f}s", "< 1.5s", "Good" if perf['total_time'] <= 1.5 else "Poor"),
            ("Download Time", f"{perf.get('download_time', 0):.3f}s", "< 1.0s", ""),
            ("Total Load Time", f"{perf['total_time']:.3f}s", "< 3.0s", "Good" if perf['total_time'] <= 3.0 else "Poor"),
            ("Page Size", f"{perf['page_size'] / 1024:.1f}KB", "< 1MB", "Good" if perf['page_size'] <= 1024000 else "Poor"),
            ("Response Code", perf['status_code'], "200", "Good" if perf['status_code'] == 200 else "Error")
        ]
        
        for i, (metric, value, benchmark, status) in enumerate(performance_data):
            ws[f'A{row + i}'] = metric
            ws[f'B{row + i}'] = value
            ws[f'C{row + i}'] = benchmark
            ws[f'D{row + i}'] = status
            
            for col in ['A', 'B', 'C', 'D']:
                cell = ws[f'{col}{row + i}']
                cell.border = border
                if i == 0:  # Header row
                    cell.font = subheader_font
                    cell.fill = subheader_fill
    
    def _create_security_sheet(self, ws, results, header_font, header_fill, subheader_font, subheader_fill, border):
        """Create security analysis sheet"""
        ws['A1'] = "Security Headers Analysis"
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        
        row = 3
        headers = [
            ("Security Header", "Present", "Description", "Impact"),
            ("X-Content-Type-Options", "✅" if results['security']['X-Content-Type-Options'] else "❌", "Prevents MIME sniffing", "Medium"),
            ("X-Frame-Options", "✅" if results['security']['X-Frame-Options'] else "❌", "Prevents clickjacking", "High"),
            ("X-XSS-Protection", "✅" if results['security']['X-XSS-Protection'] else "❌", "XSS attack protection", "Medium"),
            ("Strict-Transport-Security", "✅" if results['security']['Strict-Transport-Security'] else "❌", "Forces HTTPS", "High"),
            ("Content-Security-Policy", "✅" if results['security']['Content-Security-Policy'] else "❌", "Prevents XSS/injection", "High"),
            ("Referrer-Policy", "✅" if results['security']['Referrer-Policy'] else "❌", "Controls referrer info", "Low"),
            ("Permissions-Policy", "✅" if results['security']['Permissions-Policy'] else "❌", "Controls browser features", "Medium")
        ]
        
        for i, (header, present, desc, impact) in enumerate(headers):
            ws[f'A{row + i}'] = header
            ws[f'B{row + i}'] = present
            ws[f'C{row + i}'] = desc
            ws[f'D{row + i}'] = impact
            
            for col in ['A', 'B', 'C', 'D']:
                cell = ws[f'{col}{row + i}']
                cell.border = border
                if i == 0:  # Header row
                    cell.font = subheader_font
                    cell.fill = subheader_fill
    
    def _create_seo_sheet(self, ws, results, header_font, header_fill, subheader_font, subheader_fill, border):
        """Create SEO analysis sheet"""
        ws['A1'] = "SEO Analysis"
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        
        seo = results['seo']
        
        row = 3
        seo_data = [
            ("SEO Element", "Current Value", "Recommendation", "Status"),
            ("Page Title", seo['title'][:50] + "..." if len(seo['title']) > 50 else seo['title'], "50-60 characters", "Good" if 50 <= len(seo['title']) <= 60 else "Needs Work"),
            ("Meta Description", seo['meta_description'][:50] + "..." if len(seo['meta_description']) > 50 else seo['meta_description'] or "Missing", "150-155 characters", "Good" if 150 <= len(seo['meta_description']) <= 155 else "Needs Work"),
            ("H1 Tags Count", len(seo['h1_tags']), "Exactly 1", "Good" if len(seo['h1_tags']) == 1 else "Needs Work"),
            ("H2 Tags Count", len(seo['h2_tags']), "2-6 recommended", "Good" if 2 <= len(seo['h2_tags']) <= 6 else "Consider Adding"),
            ("Images without Alt", seo['images_without_alt'], "0 (all should have alt text)", "Good" if seo['images_without_alt'] == 0 else "Critical"),
            ("Internal Links", seo['internal_links'], "5-10 recommended", "Good" if 5 <= seo['internal_links'] <= 20 else "Consider Adding"),
            ("External Links", seo['external_links'], "2-5 recommended", "Good" if 2 <= seo['external_links'] <= 10 else "Consider Adding")
        ]
        
        for i, (element, current, recommendation, status) in enumerate(seo_data):
            ws[f'A{row + i}'] = element
            ws[f'B{row + i}'] = current
            ws[f'C{row + i}'] = recommendation
            ws[f'D{row + i}'] = status
            
            for col in ['A', 'B', 'C', 'D']:
                cell = ws[f'{col}{row + i}']
                cell.border = border
                if i == 0:  # Header row
                    cell.font = subheader_font
                    cell.fill = subheader_fill
    
    def _create_links_sheet(self, ws, results, header_font, header_fill, subheader_font, subheader_fill, border):
        """Create links analysis sheet"""
        ws['A1'] = "Links Analysis"
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        
        row = 3
        ws[f'A{row}'] = "URL"
        ws[f'B{row}'] = "Status Code"
        ws[f'C{row}'] = "Response Time (s)"
        ws[f'D{row}'] = "Status"
        ws[f'E{row}'] = "Error Message"
        
        for col in ['A', 'B', 'C', 'D', 'E']:
            cell = ws[f'{col}{row}']
            cell.font = subheader_font
            cell.fill = subheader_fill
            cell.border = border
        
        row += 1
        for link in results['links'][:100]:  # Limit to 100 links
            ws[f'A{row}'] = link['url'][:100] + "..." if len(link['url']) > 100 else link['url']
            ws[f'B{row}'] = link['status_code']
            ws[f'C{row}'] = f"{link['response_time']:.3f}"
            ws[f'D{row}'] = "Broken" if link['is_broken'] else "OK"
            ws[f'E{row}'] = link.get('error_message', '')
            
            for col in ['A', 'B', 'C', 'D', 'E']:
                cell = ws[f'{col}{row}']
                cell.border = border
                if link['is_broken']:
                    cell.fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
            row += 1
    
    def _create_images_sheet(self, ws, results, header_font, header_fill, subheader_font, subheader_fill, border):
        """Create images analysis sheet"""
        ws['A1'] = "Images Analysis"
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        
        row = 3
        ws[f'A{row}'] = "Image URL"
        ws[f'B{row}'] = "Size (KB)"
        ws[f'C{row}'] = "Format"
        ws[f'D{row}'] = "Optimized"
        ws[f'E{row}'] = "Issues"
        
        for col in ['A', 'B', 'C', 'D', 'E']:
            cell = ws[f'{col}{row}']
            cell.font = subheader_font
            cell.fill = subheader_fill
            cell.border = border
        
        row += 1
        for img in results['images'][:50]:  # Limit to 50 images
            ws[f'A{row}'] = img['url'][-50:] if len(img['url']) > 50 else img['url']
            ws[f'B{row}'] = f"{img['size_bytes'] / 1024:.1f}"
            ws[f'C{row}'] = img['format']
            ws[f'D{row}'] = "Yes" if img['is_optimized'] else "No"
            ws[f'E{row}'] = "; ".join(img['issues'])
            
            for col in ['A', 'B', 'C', 'D', 'E']:
                cell = ws[f'{col}{row}']
                cell.border = border
                if not img['is_optimized']:
                    cell.fill = PatternFill(start_color="FCF3CF", end_color="FCF3CF", fill_type="solid")
            row += 1
    
    def _create_action_items_sheet(self, ws, results, header_font, header_fill, subheader_font, subheader_fill, border):
        """Create action items sheet"""
        ws['A1'] = "Action Items & Recommendations"
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        
        row = 3
        ws[f'A{row}'] = "Priority"
        ws[f'B{row}'] = "Issue"
        ws[f'C{row}'] = "Recommendation"
        ws[f'D{row}'] = "Impact"
        
        for col in ['A', 'B', 'C', 'D']:
            cell = ws[f'{col}{row}']
            cell.font = subheader_font
            cell.fill = subheader_fill
            cell.border = border
        
        row += 1
        
        # Prioritize insights
        critical_insights = [insight for insight in results['insights'] if '⚠️ CRITICAL' in insight or 'broken links' in insight]
        warning_insights = [insight for insight in results['insights'] if insight not in critical_insights and ('⚡' in insight or '🔒' in insight)]
        other_insights = [insight for insight in results['insights'] if insight not in critical_insights and insight not in warning_insights]
        
        all_insights = [(insight, "Critical", "High") for insight in critical_insights] + \
                      [(insight, "High", "Medium") for insight in warning_insights] + \
                      [(insight, "Medium", "Low") for insight in other_insights]
        
        for insight, priority, impact in all_insights:
            ws[f'A{row}'] = priority
            ws[f'B{row}'] = insight
            ws[f'C{row}'] = self._get_recommendation(insight)
            ws[f'D{row}'] = impact
            
            for col in ['A', 'B', 'C', 'D']:
                cell = ws[f'{col}{row}']
                cell.border = border
                if priority == "Critical":
                    cell.fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
                elif priority == "High":
                    cell.fill = PatternFill(start_color="FCF3CF", end_color="FCF3CF", fill_type="solid")
            row += 1
    
    def _get_recommendation(self, insight):
        """Get specific recommendation for an insight"""
        recommendations = {
            "load time": "Optimize images, enable compression, use CDN",
            "meta description": "Add 150-155 character meta description",
            "alt text": "Add descriptive alt text to all images",
            "security headers": "Configure security headers in web server",
            "broken links": "Fix or remove broken links immediately",
            "images": "Compress images, use WebP format",
            "H1 tags": "Use exactly one H1 tag per page",
            "viewport": "Add viewport meta tag for mobile",
            "responsive": "Implement responsive image srcset"
        }
        
        for key, rec in recommendations.items():
            if key in insight.lower():
                return rec
        return "Review and optimize based on best practices"
    
    def generate_performance_suggestions(self, results: Dict) -> Dict[str, List[str]]:
        """Generate comprehensive performance improvement suggestions"""
        suggestions = {
            "🚀 Speed Optimization": [],
            "🔒 Security Enhancements": [],
            "📝 SEO Improvements": [],
            "♿ Accessibility Fixes": [],
            "📱 Mobile Optimization": [],
            "🖼️ Image Optimization": [],
            "🔗 Link Management": [],
            "💡 Technical Improvements": []
        }
        
        perf = results['performance']
        security = results['security']
        seo = results['seo']
        images = results['images']
        links = results['links']
        mobile = results['mobile']
        
        # Speed Optimization Suggestions
        if perf['total_time'] > 3.0:
            suggestions["🚀 Speed Optimization"].extend([
                "🚨 CRITICAL: Reduce server response time - consider upgrading hosting",
                "⚡ Enable gzip compression to reduce file sizes by 60-80%",
                "📦 Use a Content Delivery Network (CDN) like Cloudflare",
                "🗜️ Minify CSS, JavaScript, and HTML files",
                "⚡ Implement browser caching with appropriate cache headers"
            ])
        elif perf['total_time'] > 1.5:
            suggestions["🚀 Speed Optimization"].extend([
                "⚡ Optimize critical rendering path",
                "📦 Enable browser caching (set cache-control headers)",
                "🗜️ Minify and compress static assets",
                "⚡ Consider using HTTP/2 for faster loading"
            ])
        
        if perf['page_size'] > 2000000:  # 2MB
            suggestions["🚀 Speed Optimization"].extend([
                "📦 Reduce page size - current: {:.1f}MB (target: <1MB)".format(perf['page_size']/1024/1024),
                "🖼️ Optimize large images (compress, resize, use WebP)",
                "📜 Split large JavaScript bundles",
                "🧹 Remove unused CSS and JavaScript"
            ])
        
        # Security Enhancement Suggestions
        missing_headers = [header for header, present in security.items() if not present]
        if missing_headers:
            suggestions["🔒 Security Enhancements"].extend([
                "🛡️ Add missing security headers in your web server configuration:",
                f"   • {', '.join(missing_headers)}",
                "⚡ Quick fix for Apache: Add to .htaccess file",
                "⚡ Quick fix for Nginx: Add to server block",
                "🔧 Use online tools like securityheaders.com to test"
            ])
            
        if not security.get('Strict-Transport-Security', False):
            suggestions["🔒 Security Enhancements"].append(
                "🔐 Enable HTTPS and add HSTS header for security"
            )
        
        # SEO Improvement Suggestions
        if not seo['title']:
            suggestions["📝 SEO Improvements"].append(
                "📄 Add a compelling page title (50-60 characters)"
            )
        elif len(seo['title']) > 60:
            suggestions["📝 SEO Improvements"].append(
                f"✂️ Shorten page title from {len(seo['title'])} to 50-60 characters"
            )
        elif len(seo['title']) < 30:
            suggestions["📝 SEO Improvements"].append(
                "📏 Expand page title to 30-60 characters for better SEO"
            )
        
        if not seo['meta_description']:
            suggestions["📝 SEO Improvements"].append(
                "📄 Add meta description (150-155 characters) for better search results"
            )
        elif len(seo['meta_description']) > 155:
            suggestions["📝 SEO Improvements"].append(
                f"✂️ Shorten meta description from {len(seo['meta_description'])} to 150-155 characters"
            )
        
        if len(seo['h1_tags']) == 0:
            suggestions["📝 SEO Improvements"].append(
                "🏷️ Add exactly one H1 tag with your main keyword"
            )
        elif len(seo['h1_tags']) > 1:
            suggestions["📝 SEO Improvements"].append(
                f"🏷️ Use only one H1 tag per page (currently {len(seo['h1_tags'])})"
            )
        
        if len(seo['h2_tags']) < 2:
            suggestions["📝 SEO Improvements"].append(
                "🏷️ Add 2-6 H2 tags to structure your content better"
            )
        
        if seo['internal_links'] < 5:
            suggestions["📝 SEO Improvements"].append(
                "🔗 Add more internal links (5-10) to improve site navigation"
            )
        
        if seo['external_links'] < 2:
            suggestions["📝 SEO Improvements"].append(
                "🔗 Add 2-5 relevant external links to authoritative sources"
            )
        
        # Accessibility Fixes
        if seo['images_without_alt'] > 0:
            suggestions["♿ Accessibility Fixes"].extend([
                f"🖼️ Add alt text to {seo['images_without_alt']} images",
                "💡 Use descriptive alt text that explains the image content",
                "⚡ Quick tip: alt='Description of what's in the image'"
            ])
        
        # Mobile Optimization
        if not mobile.get('viewport_meta', False):
            suggestions["📱 Mobile Optimization"].extend([
                "📱 Add viewport meta tag: <meta name='viewport' content='width=device-width, initial-scale=1'>",
                "🔧 This is essential for mobile responsiveness"
            ])
        
        if not mobile.get('responsive_images', False):
            suggestions["📱 Mobile Optimization"].extend([
                "🖼️ Implement responsive images with srcset attribute",
                "💡 Example: <img src='image.jpg' srcset='image-small.jpg 480w, image-large.jpg 1200w'>",
                "⚡ This improves loading on mobile devices"
            ])
        
        # Image Optimization
        large_images = [img for img in images if img['size_bytes'] > 500000]
        unoptimized_images = [img for img in images if not img['is_optimized']]
        
        if large_images:
            suggestions["🖼️ Image Optimization"].extend([
                f"📦 Compress {len(large_images)} large images (>500KB)",
                "🔧 Use tools like TinyPNG, ImageOptim, or Squoosh",
                "⚡ Convert to WebP format for 25-50% size reduction",
                "📏 Resize images to actual display dimensions"
            ])
        
        if unoptimized_images:
            suggestions["🖼️ Image Optimization"].extend([
                f"🔧 Optimize {len(unoptimized_images)} images for better performance",
                "💡 Use modern formats: WebP > JPEG > PNG",
                "⚡ Enable lazy loading for images below the fold",
                "🎯 Target: <100KB per image for web"
            ])
        
        # Link Management
        broken_links = [link for link in links if link['is_broken']]
        if broken_links:
            suggestions["🔗 Link Management"].extend([
                f"🚨 Fix {len(broken_links)} broken links immediately",
                "🔧 Use tools like Screaming Frog or Broken Link Checker",
                "⚡ Set up 301 redirects for moved content",
                "🎯 Broken links hurt SEO and user experience"
            ])
        
        slow_links = [link for link in links if link['response_time'] > 3.0]
        if slow_links:
            suggestions["🔗 Link Management"].append(
                f"⚡ {len(slow_links)} links are slow (>3s) - check linked sites"
            )
        
        # Technical Improvements
        suggestions["💡 Technical Improvements"].extend([
            "📊 Set up Google Analytics and Search Console",
            "🎯 Monitor Core Web Vitals (LCP, FID, CLS)",
            "⚡ Consider implementing AMP for mobile pages",
            "🔄 Set up regular performance monitoring",
            "📈 Use tools like PageSpeed Insights weekly",
            "🧪 A/B test performance improvements",
            "📱 Test on real mobile devices, not just desktop"
        ])
        
        # Remove empty categories
        suggestions = {k: v for k, v in suggestions.items() if v}
        
        return suggestions
    
    def print_performance_suggestions(self, suggestions: Dict[str, List[str]]):
        """Print formatted performance suggestions"""
        print("\n" + "🤖" * 20 + " ULTRON PERFORMANCE BOOST GUIDE " + "🤖" * 20)
        print("🚀 Ready to supercharge your website? Here's your action plan:")
        print("=" * 80)
        
        priority_order = [
            "🚀 Speed Optimization",
            "🔒 Security Enhancements", 
            "📝 SEO Improvements",
            "♿ Accessibility Fixes",
            "📱 Mobile Optimization",
            "🖼️ Image Optimization",
            "🔗 Link Management",
            "💡 Technical Improvements"
        ]
        
        for category in priority_order:
            if category in suggestions and suggestions[category]:
                print(f"\n{category}")
                print("-" * 50)
                for i, suggestion in enumerate(suggestions[category], 1):
                    print(f"   {i}. {suggestion}")
        
        print("\n" + "💡" * 20 + " IMPLEMENTATION TIPS " + "💡" * 20)
        print("🎯 Start with Speed Optimization for immediate impact")
        print("🔒 Security fixes protect your users and boost SEO")
        print("📝 SEO improvements increase organic traffic")
        print("📱 Mobile optimization is crucial (60%+ mobile traffic)")
        print("⚡ Fix critical issues first, then work on enhancements")
        print("📊 Measure before and after to track improvements")
        
        print(f"\n🏆 Get a perfect score by implementing these suggestions!")
        print(f"🔄 Re-run Ultron after changes to track your progress")
        print("=" * 80)
    
    def generate_html_report(self, results: Dict) -> str:
        """Generate HTML report"""
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>🤖 Ultron Performance Report - {results['url']}</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                .header {{ background: #2c3e50; color: white; padding: 20px; }}
                .section {{ margin: 20px 0; padding: 15px; border: 1px solid #ddd; }}
                .insight {{ background: #f8f9fa; padding: 10px; margin: 5px 0; }}
                .score {{ font-size: 2em; font-weight: bold; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>🤖 Ultron Performance Report</h1>
                <p>{results['url']}</p>
                <p>Generated: {results['timestamp']}</p>
            </div>
            
            <div class="section">
                <h2>Performance</h2>
                <p>Load Time: {results['performance']['total_time']:.2f}s</p>
                <p>Page Size: {results['performance']['page_size'] / 1024:.1f}KB</p>
            </div>
            
            <div class="section">
                <h2>Insights</h2>
                {"".join([f'<div class="insight">{insight}</div>' for insight in results['insights']])}
            </div>
        </body>
        </html>
        """
        return html