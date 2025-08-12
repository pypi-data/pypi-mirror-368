import openpyxl
import typer
from typing import Any


def excel_row_generator(excel_path: str, sheet_name: str | None = None):
    """
    엑셀 파일을 한 줄씩 읽어 객체 형태로 반환하는 제너레이터 함수 
    """
    try:
        workbook = openpyxl.load_workbook(excel_path, read_only=True)
        
        if sheet_name:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active
        
        # 첫번째 행은 헤더 
        header = [cell.value for cell in next(sheet.iter_rows())]
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(header, row))

            yield {
                key: value.strip() if isinstance(value, str) else value
                for key, value in row_dict.items()
            }
            
    except FileNotFoundError:
        print(f"{excel_path} 경로에서 파일을 찾을 수 없습니다.")
    
    finally:
        if 'workbook' in locals():
            workbook.close()


def validate_empty_str(row: dict[str, Any], row_num: int) -> list:
    error_data: list = []
    
    # string 타입인 경우 빈 문자열 검사 
    for i, key in enumerate(row.keys()):
        if isinstance(row[key], str) and row[key].strip() == "":
            error_data.append({
                "row_num": row_num,
                "col_num": i + 1,
                "key": key
            })
    
    return error_data


def success_log(msg: str) -> None:
    typer.echo(f"[SUCCESS] {msg}", color=typer.colors.GREEN)
    

def info_log(msg: str) -> None:
    typer.echof(f"[INFO] {msg}", color=typer.colors.BRIGHT_WHITE)


def error_log(msg: str) -> None:
    typer.echo(f"[ERROR] {msg}", color=typer.colors.RED, err=True)

