from __future__ import annotations

import csv
import logging
import mimetypes
import os
from dataclasses import dataclass
from pathlib import Path
from typing import Iterator

import httpx
from httpx import URL
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

ENV_API_KEY = 'NODE_IMAGE_API_KEY'

IMAGE_EXTENSIONS = {'.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp', '.tiff', '.tif', '.svg'}


def is_image_file(file_path: str | Path) -> bool:
    return Path(file_path).suffix.lower() in IMAGE_EXTENSIONS


def get_api_key_from_env(env_key: str = ENV_API_KEY) -> str:
    return os.getenv(env_key) or ''


def is_url(url: str | URL) -> bool:
    return URL(url).scheme in ['http', 'https']


def is_path(path: str | Path) -> bool:
    return Path(path).exists()


def is_file(path: str | Path) -> bool:
    return Path(path).is_file()


def iter_files_in_path(path: str | Path) -> Iterator[Path]:
    p = Path(path)
    if not p.exists():
        return
    if p.is_file():
        yield p
    elif p.is_dir():
        for file in p.rglob('*'):
            if file.is_file():
                yield file


@dataclass
class ImageInfo:
    ext: str
    content: bytes
    content_type: str


def download_image_from_url(
    url: str,
    timeout: int = 10,
    logger: logging.Logger | None = None,
) -> ImageInfo:
    if not is_url(url):
        raise ValueError(f'Invalid URL: {url}')

    logger = logger or logging.getLogger(__name__)
    logger.info(f'Downloading image from URL: {url}')

    response = httpx.get(url, timeout=timeout, follow_redirects=True)
    logger.debug(f'Response status code: {response.status_code}, headers: {response.headers}, content: {response.text}')
    if response.status_code != 200:
        message = (
            f'Failed to download image from URL: {url}, status code: {response.status_code}, body: {response.text}'
        )
        logger.error(message)
        raise ValueError(message)

    content_type = response.headers.get('Content-Type')
    if not content_type:
        content_type = 'image/jpeg'

    ext = mimetypes.guess_extension(content_type)
    if not ext:
        ext = '.jpg'

    return ImageInfo(ext=ext, content=response.content, content_type=content_type)


def write_csv_file(file_path: str, field_order: list[str], data_rows: list[dict]) -> None:
    with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=field_order)
        writer.writeheader()
        for row in data_rows:
            # 确保所有字段都有值，缺失的用空字符串填充
            complete_row = {field: row.get(field, '') for field in field_order}
            writer.writerow(complete_row)


def write_xlsx_file(file_path: str, field_order: list[str], data_rows: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = 'ImageInfos'

    # 写入表头
    for col, field in enumerate(field_order, 1):
        ws.cell(row=1, column=col, value=field)

    # 写入数据
    for row_idx, row_data in enumerate(data_rows, 2):
        for col_idx, field in enumerate(field_order, 1):
            value = row_data.get(field, '')
            ws.cell(row=row_idx, column=col_idx, value=value)

    # 自动调整列宽
    for col in range(1, len(field_order) + 1):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].auto_size = True

    wb.save(file_path)
