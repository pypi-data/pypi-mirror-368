"""Tests for nodeimage._utils module."""

from __future__ import annotations

import csv
import os
from pathlib import Path
from unittest.mock import Mock, patch

import pytest
from httpx import URL
from openpyxl import load_workbook

from nodeimage._utils import (
    ENV_API_KEY,
    IMAGE_EXTENSIONS,
    ImageInfo,
    download_image_from_url,
    get_api_key_from_env,
    is_file,
    is_image_file,
    is_path,
    is_url,
    iter_files_in_path,
    write_csv_file,
    write_xlsx_file,
)


class TestBasicUtilities:
    """Test basic utility functions."""

    def test_is_image_file_valid_extensions(self):
        """Test is_image_file with valid image extensions."""
        for ext in IMAGE_EXTENSIONS:
            assert is_image_file(f'test{ext}')
            assert is_image_file(f'test{ext.upper()}')  # Test case insensitive

    def test_is_image_file_invalid_extensions(self):
        """Test is_image_file with invalid extensions."""
        invalid_files = ['test.txt', 'test.pdf', 'test.doc', 'test']
        for file_path in invalid_files:
            assert not is_image_file(file_path)

    def test_is_image_file_with_path_object(self):
        """Test is_image_file with Path objects."""
        assert is_image_file(Path('test.jpg'))
        assert not is_image_file(Path('test.txt'))

    @patch.dict(os.environ, {ENV_API_KEY: 'test_api_key'})
    def test_get_api_key_from_env_exists(self):
        """Test get_api_key_from_env when key exists."""
        assert get_api_key_from_env() == 'test_api_key'

    @patch.dict(os.environ, {}, clear=True)
    def test_get_api_key_from_env_not_exists(self):
        """Test get_api_key_from_env when key doesn't exist."""
        assert get_api_key_from_env() == ''

    @patch.dict(os.environ, {'CUSTOM_KEY': 'custom_value'})
    def test_get_api_key_from_env_custom_key(self):
        """Test get_api_key_from_env with custom key."""
        assert get_api_key_from_env('CUSTOM_KEY') == 'custom_value'

    def test_is_url_valid_http(self):
        """Test is_url with valid HTTP URLs."""
        valid_urls = [
            'http://example.com',
            'https://example.com',
            'http://example.com/path',
            'https://example.com/path?query=1',
        ]
        for url in valid_urls:
            assert is_url(url)

    def test_is_url_invalid(self):
        """Test is_url with invalid URLs."""
        invalid_urls = [
            'ftp://example.com',
            'file:///path/to/file',
            '/local/path',
            'example.com',
            '',
        ]
        for url in invalid_urls:
            assert not is_url(url)

    def test_is_url_with_url_object(self):
        """Test is_url with URL objects."""
        assert is_url(URL('https://example.com'))
        assert not is_url(URL('ftp://example.com'))


class TestFileOperations:
    """Test file operation utilities."""

    def test_is_path_exists(self, tmp_path):
        """Test is_path with existing path."""
        test_file = tmp_path / 'test.txt'
        test_file.write_text('test')
        
        assert is_path(test_file)
        assert is_path(str(test_file))
        assert is_path(tmp_path)

    def test_is_path_not_exists(self):
        """Test is_path with non-existing path."""
        assert not is_path('/non/existing/path')
        assert not is_path(Path('/non/existing/path'))

    def test_is_file_valid(self, tmp_path):
        """Test is_file with valid file."""
        test_file = tmp_path / 'test.txt'
        test_file.write_text('test')
        
        assert is_file(test_file)
        assert is_file(str(test_file))

    def test_is_file_directory(self, tmp_path):
        """Test is_file with directory."""
        assert not is_file(tmp_path)

    def test_is_file_not_exists(self):
        """Test is_file with non-existing file."""
        assert not is_file('/non/existing/file.txt')

    def test_iter_files_in_path_single_file(self, tmp_path):
        """Test iter_files_in_path with single file."""
        test_file = tmp_path / 'test.txt'
        test_file.write_text('test')
        
        files = list(iter_files_in_path(test_file))
        assert len(files) == 1
        assert files[0] == test_file

    def test_iter_files_in_path_directory(self, tmp_path):
        """Test iter_files_in_path with directory."""
        # Create test files
        (tmp_path / 'file1.txt').write_text('test1')
        (tmp_path / 'file2.jpg').write_text('test2')
        subdir = tmp_path / 'subdir'
        subdir.mkdir()
        (subdir / 'file3.png').write_text('test3')
        
        files = list(iter_files_in_path(tmp_path))
        file_names = {f.name for f in files}
        
        assert len(files) == 3
        assert file_names == {'file1.txt', 'file2.jpg', 'file3.png'}

    def test_iter_files_in_path_not_exists(self):
        """Test iter_files_in_path with non-existing path."""
        files = list(iter_files_in_path('/non/existing/path'))
        assert len(files) == 0


class TestImageInfo:
    """Test ImageInfo dataclass."""

    def test_image_info_creation(self):
        """Test ImageInfo creation."""
        info = ImageInfo(
            ext='.jpg',
            content=b'fake_image_data',
            content_type='image/jpeg'
        )
        
        assert info.ext == '.jpg'
        assert info.content == b'fake_image_data'
        assert info.content_type == 'image/jpeg'


class TestDownloadImage:
    """Test image download functionality."""

    @patch('nodeimage._utils.httpx.get')
    def test_download_image_from_url_success(self, mock_get):
        """Test successful image download."""
        # Mock successful response
        mock_response = Mock()
        mock_response.status_code = 200
        mock_response.content = b'fake_image_data'
        mock_response.headers = {'Content-Type': 'image/jpeg'}
        mock_get.return_value = mock_response

        result = download_image_from_url('https://example.com/image.jpg')
        
        assert isinstance(result, ImageInfo)
        assert result.ext == '.jpg'  # mimetypes.guess_extension('image/jpeg') returns '.jpg'
        assert result.content == b'fake_image_data'
        assert result.content_type == 'image/jpeg'

    @patch('nodeimage._utils.httpx.get')
    def test_download_image_from_url_no_content_type(self, mock_get):
        """Test image download without Content-Type header."""
        mock_response = Mock()
        mock_response.status_code = 200
        mock_response.content = b'fake_image_data'
        mock_response.headers = {}
        mock_get.return_value = mock_response

        result = download_image_from_url('https://example.com/image.jpg')
        
        assert result.content_type == 'image/jpeg'
        assert result.ext == '.jpg'

    @patch('nodeimage._utils.httpx.get')
    def test_download_image_from_url_failure(self, mock_get):
        """Test failed image download."""
        mock_response = Mock()
        mock_response.status_code = 404
        mock_response.text = 'Not Found'
        mock_get.return_value = mock_response

        with pytest.raises(ValueError, match='Failed to download image'):
            download_image_from_url('https://example.com/nonexistent.jpg')

    def test_download_image_from_url_invalid_url(self):
        """Test download with invalid URL."""
        with pytest.raises(ValueError, match='Invalid URL'):
            download_image_from_url('not_a_url')


class TestFileExport:
    """Test file export functionality."""

    def test_write_csv_file(self, tmp_path):
        """Test CSV file writing."""
        csv_file = tmp_path / 'test.csv'
        field_order = ['name', 'age', 'city']
        data_rows = [
            {'name': 'Alice', 'age': '25', 'city': 'New York'},
            {'name': 'Bob', 'age': '30'},  # Missing city field
            {'name': 'Charlie', 'age': '35', 'city': 'London', 'extra': 'ignored'},
        ]

        write_csv_file(str(csv_file), field_order, data_rows)

        # Verify file was created and content is correct
        assert csv_file.exists()
        
        with open(csv_file, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            rows = list(reader)
            
        assert len(rows) == 3
        assert rows[0] == {'name': 'Alice', 'age': '25', 'city': 'New York'}
        assert rows[1] == {'name': 'Bob', 'age': '30', 'city': ''}  # Missing field filled with empty string
        assert rows[2] == {'name': 'Charlie', 'age': '35', 'city': 'London'}

    def test_write_xlsx_file(self, tmp_path):
        """Test XLSX file writing."""
        xlsx_file = tmp_path / 'test.xlsx'
        field_order = ['name', 'age', 'city']
        data_rows = [
            {'name': 'Alice', 'age': '25', 'city': 'New York'},
            {'name': 'Bob', 'age': '30'},  # Missing city field
            {'name': 'Charlie', 'age': '35', 'city': 'London'},
        ]

        write_xlsx_file(str(xlsx_file), field_order, data_rows)

        # Verify file was created and content is correct
        assert xlsx_file.exists()
        
        wb = load_workbook(xlsx_file)
        ws = wb.active
        
        # Check headers
        assert ws.cell(row=1, column=1).value == 'name'
        assert ws.cell(row=1, column=2).value == 'age'
        assert ws.cell(row=1, column=3).value == 'city'
        
        # Check data
        assert ws.cell(row=2, column=1).value == 'Alice'
        assert ws.cell(row=2, column=2).value == '25'
        assert ws.cell(row=2, column=3).value == 'New York'
        
        assert ws.cell(row=3, column=1).value == 'Bob'
        assert ws.cell(row=3, column=2).value == '30'
        # openpyxl treats empty string as None
        assert ws.cell(row=3, column=3).value in ('', None)  # Missing field
        
        assert ws.cell(row=4, column=1).value == 'Charlie'
        assert ws.cell(row=4, column=2).value == '35'
        assert ws.cell(row=4, column=3).value == 'London'

    def test_write_csv_file_empty_data(self, tmp_path):
        """Test CSV file writing with empty data."""
        csv_file = tmp_path / 'empty.csv'
        field_order = ['name', 'age']
        data_rows = []

        write_csv_file(str(csv_file), field_order, data_rows)

        assert csv_file.exists()
        
        with open(csv_file, 'r', encoding='utf-8') as f:
            content = f.read()
            
        # Should only contain headers
        assert content.strip() == 'name,age'

    def test_write_xlsx_file_empty_data(self, tmp_path):
        """Test XLSX file writing with empty data."""
        xlsx_file = tmp_path / 'empty.xlsx'
        field_order = ['name', 'age']
        data_rows = []

        write_xlsx_file(str(xlsx_file), field_order, data_rows)

        assert xlsx_file.exists()
        
        wb = load_workbook(xlsx_file)
        ws = wb.active
        
        # Should only contain headers
        assert ws.cell(row=1, column=1).value == 'name'
        assert ws.cell(row=1, column=2).value == 'age'
        assert ws.cell(row=2, column=1).value is None  # No data rows
