#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
List report generation module for ddQuint with tabular format and buffer zone support.

Generates Excel reports in list format with chromosome data organized in columns.
Includes copy number highlighting for aneuploidies and buffer zones with proper
column-first well sorting.
"""

import os
import logging
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

from ..config import Config, ReportGenerationError

logger = logging.getLogger(__name__)


def create_list_report(results, output_path):
    """
    Create a list-format Excel report with chromosome data in columns.
    
    Args:
        results (list): List of result dictionaries for each well
        output_path (str): Path to save the Excel report
        
    Returns:
        str: Path to the saved Excel report
        
    Raises:
        ReportGenerationError: If report creation fails
        ValueError: If results data is invalid
    """
    config = Config.get_instance()
    
    if not results:
        error_msg = "No results provided for list report generation"
        logger.error(error_msg)
        raise ValueError(error_msg)
    
    # Filter results by minimum usable droplets for successful analysis
    # but keep error/insufficient data results for visibility
    filtered_results = []
    excluded_count = 0
    
    for result in results:
        usable_droplets = result.get('usable_droplets', 0)
        has_error = result.get('error') is not None
        
        # Include wells with sufficient droplets OR wells with errors/insufficient data
        if usable_droplets >= config.MIN_USABLE_DROPLETS or has_error or usable_droplets > 0:
            filtered_results.append(result)
        else:
            # Only exclude wells with absolutely no data
            excluded_count += 1
            well_id = result.get('well', 'Unknown')
            logger.debug(f"Excluding well {well_id} from report: no data available")
    
    if excluded_count > 0:
        logger.debug(f"Excluded {excluded_count} wells from analysis report due to no data")
    
    if not filtered_results:
        error_msg = f"No wells with any data available for reporting"
        logger.warning(error_msg)
        # Still create an empty report rather than failing
        filtered_results = []
    
    # Get the number of chromosomes from config
    chromosome_keys = config.get_chromosome_keys()
    num_chromosomes = len(chromosome_keys)
    
    logger.debug(f"Creating list report for {len(filtered_results)} results with {num_chromosomes} chromosomes")
    logger.debug(f"Output path: {output_path}")
    
    try:
        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "List Results"
        logger.debug("Created new workbook")
        
        # Sort results by well ID in column-first order
        sorted_results = sorted(filtered_results, key=lambda x: parse_well_id_column_first(x.get('well', '')))
        
        # Set up headers with proper structure
        setup_headers(ws, chromosome_keys)
        
        # Fill in data for each well
        fill_well_data(ws, sorted_results, chromosome_keys, config)
        
        # Apply formatting
        apply_formatting(ws, len(sorted_results), chromosome_keys)
        
        # Save the workbook
        wb.save(output_path)
        logger.debug(f"List report saved successfully to {output_path}")
        return output_path
        
    except Exception as e:
        error_msg = f"Error creating list report for {os.path.basename(output_path)}: {str(e)}"
        logger.error(error_msg)
        logger.debug(f"Error details: {str(e)}", exc_info=True)
        raise ReportGenerationError(error_msg) from e

def parse_well_id_column_first(well_id):
    """
    Parse well ID to support column-first sorting.
    
    Args:
        well_id (str): Well identifier like 'A01', 'B12', etc.
        
    Returns:
        tuple: (column_number, row_number) for column-first sorting
    """
    if not well_id:
        return (999, 999)  # Put empty well IDs at the end
    
    # Extract letter(s) for row and number(s) for column
    row_part = ''
    col_part = ''
    
    for char in well_id:
        if char.isalpha():
            row_part += char
        elif char.isdigit():
            col_part += char
    
    # Convert row letters to number (A=1, B=2, etc.)
    if row_part:
        row_number = 0
        for i, char in enumerate(reversed(row_part.upper())):
            row_number += (ord(char) - ord('A') + 1) * (26 ** i)
    else:
        row_number = 999  # Put malformed wells at the end
    
    # Convert column to integer
    try:
        col_number = int(col_part) if col_part else 0
    except ValueError:
        col_number = 999
    
    # Return (column, row) for column-first sorting
    return (col_number, row_number)


def setup_headers(ws, chromosome_keys):
    """
    Set up headers with proper merging and structure.
    
    Args:
        ws: Worksheet to modify
        chromosome_keys (list): List of chromosome identifiers
    """
    num_chromosomes = len(chromosome_keys)

    # Row 1: Main headers
    ws.cell(row=1, column=1, value="Well")
    ws.cell(row=2, column=1, value="")

    ws.cell(row=1, column=2, value="Sample")
    ws.cell(row=2, column=2, value="")

    # Relative Copy Number section
    rel_start = 3
    rel_end = 2 + num_chromosomes
    ws.cell(row=1, column=rel_start, value="Relative Copy Number")
    for i in range(rel_start, rel_end + 1):
        ws.cell(row=1, column=i + 1, value="")

    # Droplet Readouts section with 3 columns
    droplet_start = rel_end + 1
    droplet_end = droplet_start + 3 - 1
    ws.cell(row=1, column=droplet_start, value="Droplet Readouts")
    # Clear merged cells to prevent unwanted values
    for i in range(droplet_start, droplet_end + 1):
        ws.cell(row=1, column=i + 1, value="")

    # Second row: Droplet readout column labels
    droplet_subcols = ["Usable", "Positive", "Overall"]
    for i, label in enumerate(droplet_subcols):
        ws.cell(row=2, column=droplet_start + i, value=label)

    # Absolute Copy Number section
    abs_start = droplet_end + 1
    abs_end = abs_start + num_chromosomes - 1
    ws.cell(row=1, column=abs_start, value="Positive Droplet Count")
    for i in range(abs_start, abs_end + 1):
        ws.cell(row=1, column=i + 1, value="")

    # Merge cells for headers
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
    ws.merge_cells(start_row=1, start_column=rel_start, end_row=1, end_column=rel_end)
    ws.merge_cells(start_row=1, start_column=droplet_start, end_row=1, end_column=droplet_end)
    ws.merge_cells(start_row=1, start_column=abs_start, end_row=1, end_column=abs_end)

    # Row 2: Chromosome headers for Relative and Absolute Copy Number
    for i, chrom_key in enumerate(chromosome_keys):
        chrom_label = f"Chr{chrom_key.replace('Chrom', '')}"
        ws.cell(row=2, column=rel_start + i, value=chrom_label)
        ws.cell(row=2, column=abs_start + i, value=chrom_label)

    # Apply formatting to headers
    for cell in ws[1]:
        if cell.value:
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal='center', vertical='center')

    for cell in ws[2]:
        if cell.value:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')

def fill_well_data(ws, sorted_results, chromosome_keys, config):
    """
    Fill in data for each well with buffer zone and aneuploidy highlighting.
    
    Args:
        ws: Worksheet to modify
        sorted_results (list): Sorted list of result dictionaries
        chromosome_keys (list): List of chromosome identifiers
        config: Configuration instance
    """
    num_chromosomes = len(chromosome_keys)
    rel_start = 3
    rel_end = rel_start + num_chromosomes - 1
    
    droplet_start = rel_end + 1
    droplet_end = droplet_start + 2  # 3 columns for droplets
    
    abs_start = droplet_end + 1

    for row_idx, result in enumerate(sorted_results, start=3):
        well_id = result.get('well', '')
        
        # Well ID
        well_cell = ws.cell(row=row_idx, column=1, value=well_id)
        well_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Sample name
        sample_name = result.get('sample_name', '')
        if not sample_name:
            filename = result.get('filename', '')
            sample_name = os.path.splitext(filename)[0] if filename else well_id
        
        sample_cell = ws.cell(row=row_idx, column=2, value=sample_name)
        sample_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Get data for this well
        counts = result.get('counts', {})
        copy_numbers = result.get('copy_numbers', {})
        copy_number_states = result.get('copy_number_states', {})
        has_aneuploidy = result.get('has_aneuploidy', False)
        has_buffer_zone = result.get('has_buffer_zone', False)
        has_error = result.get('error') is not None
        
        # Extract droplet metrics from clustering results
        total_droplets = result.get('total_droplets', 0)
        usable_droplets = result.get('usable_droplets', 0) 
        negative_droplets = result.get('negative_droplets', 0)
        
        # Calculate positive droplets: Overall - Negative
        positive_droplets = total_droplets - negative_droplets
        
        # Determine row fill based on status
        row_fill = None
        if has_buffer_zone:
            row_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        elif has_aneuploidy:
            row_fill = PatternFill(start_color="F2CEEF", end_color="F2CEEF", fill_type="solid")
        elif has_error:
            row_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")  # Light red for errors
        
        if row_fill:
            well_cell.fill = row_fill
            sample_cell.fill = row_fill
        
        # Fill relative and absolute copy number data
        _fill_chromosome_data(ws, row_idx, rel_start, abs_start, chromosome_keys, 
                             copy_numbers, counts, copy_number_states, row_fill, 
                             has_buffer_zone, has_aneuploidy, has_error)
        
        # Populate droplet readout columns with calculated values
        droplet_values = [usable_droplets, positive_droplets, total_droplets]
        for i, val in enumerate(droplet_values):
            cell = ws.cell(row=row_idx, column=droplet_start + i, value=val if val > 0 else "")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if row_fill:
                cell.fill = row_fill


def _fill_chromosome_data(ws, row_idx, rel_start, abs_start, chromosome_keys, 
                         copy_numbers, counts, copy_number_states, row_fill, 
                         has_buffer_zone, has_aneuploidy, has_error):
    """
    Fill chromosome data with appropriate highlighting.
    
    Args:
        ws: Worksheet to modify
        row_idx (int): Current row index
        rel_start (int): Starting column for relative data
        abs_start (int): Starting column for absolute data
        chromosome_keys (list): List of chromosome identifiers
        copy_numbers (dict): Copy number values
        counts (dict): Absolute count values
        copy_number_states (dict): Copy number state classifications
        row_fill: PatternFill for row-level highlighting
        has_buffer_zone (bool): Whether sample has buffer zone
        has_aneuploidy (bool): Whether sample has aneuploidy
        has_error (bool): Whether sample has error
    """
    for i, chrom_key in enumerate(chromosome_keys):
        # Relative copy numbers
        rel_cell = ws.cell(row=row_idx, column=rel_start + i)
        rel_count = copy_numbers.get(chrom_key)
        # Show copy numbers even for error cases if they exist (partial analysis)
        if rel_count is not None:
            rel_cell.value = round(rel_count, 2)
            rel_cell.number_format = '0.00'
        else:
            rel_cell.value = ""
        rel_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Absolute copy numbers
        abs_cell = ws.cell(row=row_idx, column=abs_start + i)
        abs_count = counts.get(chrom_key, 0)
        # Show counts even for error cases if they exist (partial analysis)
        if abs_count > 0:
            abs_cell.value = abs_count
        else:
            abs_cell.value = ""
        abs_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Apply highlighting
        if has_buffer_zone:
            # Buffer zone samples get uniform dark grey fill
            rel_cell.fill = row_fill
            abs_cell.fill = row_fill
        elif has_aneuploidy and not has_error:
            chrom_state = copy_number_states.get(chrom_key, 'euploid')
            if chrom_state == 'aneuploidy':
                # Individual chromosome aneuploidy highlighting
                chrom_fill = PatternFill(start_color="D86DCD", end_color="D86DCD", fill_type="solid")
                rel_cell.fill = chrom_fill
                abs_cell.fill = chrom_fill
            else:
                # Non-aneuploidy chromosome in aneuploidy sample
                rel_cell.fill = row_fill
                abs_cell.fill = row_fill
        elif row_fill:
            # Error cases or other row-level highlighting
            rel_cell.fill = row_fill
            abs_cell.fill = row_fill


def apply_formatting(ws, num_results, chromosome_keys):
    """
    Apply borders, column widths, and freeze panes.
    
    Args:
        ws: Worksheet to modify
        num_results (int): Number of result rows
        chromosome_keys (list): List of chromosome identifiers
    """
    num_chromosomes = len(chromosome_keys)
    rel_start = 3
    droplets_end = rel_start + num_chromosomes + 2
    abs_start = rel_start + num_chromosomes + 3
    max_col = abs_start + num_chromosomes - 1
    max_row = num_results + 2
    
    # Border styles
    thick = Side(style='medium')
    thin = Side(style='thin')
    
    # Apply borders
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            
            # Initialize borders
            left_border = None
            right_border = None
            top_border = None
            bottom_border = None
            
            # Thick borders for table outline
            if row == 1:
                top_border = thick
            if row == max_row:
                bottom_border = thick
            if col == 1:
                left_border = thick
            if col == max_col:
                right_border = thick
            
            # Thin borders around sections
            if col == rel_start and row >= 1:
                left_border = thin
            if col == rel_start + num_chromosomes - 1 and row >= 1:
                right_border = thin
            if col == droplets_end and row >= 1:
                right_border = thin
                
            # Thick border at bottom of header row
            if row == 2:
                bottom_border = thick
                
            # Apply border if at least one side has a style
            if any([left_border, right_border, top_border, bottom_border]):
                cell.border = Border(left=left_border, right=right_border, 
                                   top=top_border, bottom=bottom_border)
    
    # Set column widths
    ws.column_dimensions['A'].width = 8  # Well
    ws.column_dimensions['B'].width = 15  # Sample
    
    # Set widths for chromosome columns
    for col in range(rel_start, max_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 8
    
    # Freeze panes (freeze top 2 rows and first 2 columns)
    ws.freeze_panes = ws.cell(row=3, column=3)