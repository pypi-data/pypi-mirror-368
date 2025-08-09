# -*- coding: utf-8 -*-
"""
Created on Mon Dec 12 12:22:27 2022

@author: Franco, Bruno Agustín 

Module depeloped to generate the output results from DP4+App. 
It generate the final file for calculations, append new theory levels to the 
data base when parametrization is used and edits appareance of the spreedsheets. 
"""

#import os 
import pandas as pd
import webbrowser, subprocess, os

from time import gmtime, strftime
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.numbers import FORMAT_PERCENTAGE
from openpyxl.styles.colors import Color
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter, column_index_from_string

#### ADD RELATIVE IMPORTS
from . import bugs_a_warning_module as warn

    #----------------------------------------------------------------------- 
def gen_out_dp4(working_folder, mode, isoms_list, solvent, stand, the_lev, 
                thelev_warns, G09command, parameters,  outputs, energy_command = None): 
    '''generate output Excel file of a correlation calc
    Uses the input and output dictionaries for the construction
    At the end, adds signature and embellish the results using auxiliar functions
    '''    
    output_file = f'{mode}_DP4+_results.xlsx'
    
    output_file = os.path.join(working_folder,output_file)
    
    i = 1
    while os.path.exists(output_file):
        output_file = output_file[:-5] + f'_{i}'+output_file[-5:]
        i +=1
    
    del i
    
    output_file = os.path.join(working_folder,output_file)
        
    outputs['results'] = rearrange_results (outputs['results'])
    
    with pd.ExcelWriter(output_file) as writer: 
    
        for sheet, matrix in outputs.items(): 
            if 'exp' in sheet : 
                exp_highlights = warn.exp_data_control(matrix)
                continue
            
            if 'e_sca' in sheet: 
                e_highlights = warn.sca_e_control(outputs['exp'], matrix)
            
            try: 
                matrix.to_excel(writer,sheet_name=sheet, 
                                index=True,
                                float_format="%.4f", 
                                startrow=2)
            except: 
                matrix = pd.DataFrame(matrix,
                                      columns=isoms_list,
                                      index=outputs['exp'].index)
                matrix = pd.concat([outputs['exp'],matrix], axis=1)
                
                matrix.to_excel(writer,
                                sheet_name=sheet, 
                                index=False,
                                float_format="%.10f")
        
        stand.to_excel(writer,sheet_name='parameters', 
                                 index=True, startrow=0, header= True)

        parameters.to_excel(writer,sheet_name='parameters', 
                            index=True,float_format="%.3f", startrow=12)

    add_signs (output_file, 'results', mode , the_lev, 
               solvent, G09command, thelev_warns, energy_command=energy_command)
                
    edit_appearance(output_file, 'results', len(isoms_list))
    
    add_highlights_warn(output_file, 'e_sca',
                        exp_highlights,
                        e_highlights, 
                        HALO= True if'HALO' in mode else False )
      
    try: 
        subprocess.run(['open', output_file])
    except:
        webbrowser.open_new(output_file)
    
    return output_file

def add_signs(file, sheet, mode, the_lev, solvent, nmr_command, 
              thelev_warns, energy_command = None): 
    '''Adds signatures and sign to a given sheet in the .xlsx file
    It's design to embellish the 'results' sheet according the position 
    of its outputs. However, it's easy to modify for other presentations.
    
    Adds the signature in the first cell and the developer icon.
    At the end, adds labels of the theory level used in the calcs and warns 
    about G09 matching with the input. 
    '''
    wb = load_workbook(filename = file)
    ws = wb[sheet]
    
    ws.cell(1 , 1).value = f'DP4+ {mode}'
    
    ws.cell(14 , 1).value = 'Cite this: J. Nat. Prod. 2023, 86, 10, 2360–2367 - '    
    ws.cell(14 , 4).hyperlink = 'https://doi.org/10.1021/acs.jnatprod.3c00566'
    
    print (mode)
    
    if 'gibbs' in energy_command : 
        ws.cell(15 , 1).value = '                  Magnetic Resonance in Chemistry, 63(1), 74-85 - '    
        ws.cell(15 , 5).hyperlink = ' https://doi.org/10.1002/mrc.5491'
    
    if 'HALO' in mode : 
        ws.cell(15 , 1).value = '                  Waiting for publication - '    
        # ws.cell(16 , 4).hyperlink = ' https://doi.org/10.1002/mrc.5491'
    
    #row = ws.max_row +3  #startrow
    row = 18  #startrow
    
    if 'Custom' not in mode: 
        the_lev = the_lev.split('.')
        the_lev = '/'.join(the_lev)
        
    ws.cell(row , 1).value = f'Theory Level selected : {the_lev} ({solvent})'
    ws.cell(row+1 , 1).value = f'NMR command lines: {nmr_command}'
    if energy_command: 
        ws.cell(row+2 , 1).value = f'ENERGY command lines: {energy_command}'

    if 'Custom' in mode: 
        ws.cell(row+4 , 1).value = thelev_warns.pop()
        
    elif thelev_warns: 
        ws.cell(row+4 , 1).value ='¡Warning! Calcs and Theory Level do not match'
        ws.cell(row+5, 1).value ='These are the inconsistencies:'
        
        # print (inputs['warn'])
        for i,j in enumerate(thelev_warns):
            ws.cell(row+6+i , 1).value = '   x: '+ j 
            
        # ws.cell(row+5, 1).value = inputs['warn']
    else: 
        ws.cell(row+4 , 1).value =u'Theory Level: OK \u2713'
    
    wb.save(file)
    
    return

def edit_appearance(file, sheet, cant_isom):
    '''Edits the appearance of the given sheet in the .xlsx file
    It's design to embellish the 'results' sheet according the position 
    of its outputs. However, it's easy to modify for other presentations. 
    '''
    wb = load_workbook(file)
    ws = wb[sheet]
    ws.row_dimensions[1].height = 32
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[13].height = 9
    ws.row_dimensions[17].height = 9
    ws.column_dimensions['A'].width = 26.5
    for i in range(4, 7):
        ws[f'A{i}'].fill = PatternFill(start_color = 'BFC9CA', end_color = 'BFC9CA', fill_type = "solid")
    for i in range(7, 10):
        ws[f'A{i}'].fill = PatternFill(start_color = 'E5E7E9', end_color = 'E5E7E9', fill_type = "solid")
    for i in range(10, 13):
        ws[f'A{i}'].fill = PatternFill(start_color = '85C1E9', end_color = '85C1E9', fill_type = "solid")
    
    for row in range(4,13):
        for col in range(cant_isom) :
            ws[f'{chr(66+col)}{row}'].number_format = FORMAT_PERCENTAGE
        
    for col in ['B','C','D']:
        ws[f'{col}25'].font = Font(name='Symbol')
    
    ws['A1'].font = Font(name='Lucida Sans', size=14, bold=True,
                         italic=False, vertAlign=None, underline='none',
                         strike=False, color='FF000000')
    
    for cel in ['A14','A15']:
        ws[cel].font = Font(name='Calibri', size=11, bold=False,
                             italic=True, vertAlign=None, underline='none',
                             strike=False, color='FF000000')
    
    for cel in ['D14','D15', 'E15']:
        ws[cel].font = Font(name='Calibri', size=11, bold=False,
                             italic=True, vertAlign=None, underline='single',
                             strike=False, color='0000FF')
    
    for cel in ['A21', 'A22', 'A23', 'A24', 'A25', 'A26']: 
        ws[cel].font = Font(name='Calibri', size=10, bold=False,
                             italic=False, vertAlign=None, underline='none',
                             strike=False, color='FF000000')
    
    
    #remember to: pip install Pillow (or add to dependencies)
    
    img1 = os.path.join(os.path.dirname(__file__) , 'GUI2' ,'logo_CONICET.png')
    img1 = Image(img1)
    img1.height = img1.height/4
    img1.width= img1.width/4
    img1.anchor = 'E1'
    ws.add_image(img1)

    img2 = os.path.join(os.path.dirname(__file__) , 'GUI2' ,'logo_INGEBIO.png')
    img2 = Image(img2)
    img2.height = img2.height/2
    img2.width= img2.width/2
    img2.anchor = 'C1'
    ws.add_image(img2)
    
    ws.sheet_view.showGridLines = False

    ws = wb['parameters']
    for col in ['B','C','D']:
        ws[f'{col}13'].font = Font(name='Symbol', bold=True)

    wb.save(file)
    return
#-----------------------------------------------------------------------    
def gen_out_custom(mode, name, C_TMS, H_TMS, param):
    '''It appends the parameters obtained by Reparametrization to the data
    base of the programme. It follows some rules to order the data correctly
    so as it can be use later in DP4+Custom calcs.
    '''  
    data_base = os.path.join(os.path.dirname(__file__), 'data_base_Custom.xlsx')

    wb = load_workbook(filename = data_base)
    ws = wb['standard']
    levels = wb.sheetnames
    
    if name in levels: 
        for i in range (1, ws.max_row+1): 
            if name == ws.cell(i , 1).value :
                row = i
                wb.remove(wb[name])
                break
    else : 
        row = ws.max_row +1  #startrow
    
    ws.cell(row , 1).value = name
    ws.cell(row , 2).value = C_TMS
    ws.cell(row , 3).value = H_TMS
    ws.cell(row , 4).value = strftime("%m/%d/%Y", gmtime())
    ws.cell(row , 5).value = mode

    wb.save(data_base)
    
    with pd.ExcelWriter(data_base, engine = 'openpyxl', mode='a') as writer: 
        param.to_excel(writer, sheet_name= name)
     
    return  

def save_MSTD_HALO(mode, name, HALO_stand):
    ''' Save HALO tensors for standards
    '''  
    data_base = os.path.join(os.path.dirname(__file__), 'MSTD-Custom-Stand.xlsx')

    wb = load_workbook(filename = data_base)
    ws = wb['Hoja 1']
    levels = wb.sheetnames
    
    if name in levels: 
        for i in range (1, ws.max_row+1): 
            if name == ws.cell(i , 1).value :
                row = i
                wb.remove(wb[name])
                break
    else : 
        row = ws.max_row +1  #startrow
    
    ws.cell(row , 1).value = name
    ws.cell(row , 2).value = HALO_stand['Cl_sp2'][0]
    ws.cell(row , 3).value = HALO_stand['Cl_sp3'][0]
    ws.cell(row , 4).value = HALO_stand['Br_sp2'][0]
    ws.cell(row , 5).value = HALO_stand['Br_sp3'][0]
    
    ws.cell(row , 6).value = HALO_stand['Cl_sp2'][1]
    ws.cell(row , 7).value = HALO_stand['Cl_sp3'][1]
    ws.cell(row , 8).value = HALO_stand['Br_sp2'][1]
    ws.cell(row , 9).value = HALO_stand['Br_sp3'][1]
    
    
    # ws.cell(row , 10).value = strftime("%m/%d/%Y", gmtime())
    # ws.cell(row , 10).value = mode

    wb.save(data_base)
     
    return 
    
def add_highlights_warn(file, sheet, exp_hl, e_hl, HALO = False):
    ''' Given a file, a sheet from it and two lists of Excel coordinates, it
    highlights/paint the given cells to indicate some warning presence. 
    '''
    wb = load_workbook(file)
    ws = wb[sheet]
    
    for cell in exp_hl + e_hl :
        
        if HALO: 
            column = ''.join(filter(str.isalpha, cell))  # Obtiene la parte alfabética (columna)
            row = ''.join(filter(str.isdigit, cell))    # Obtiene la parte numérica (fila)
            new_column_index = column_index_from_string(column) + 1
            cell = f"{get_column_letter(new_column_index)}{row}"
        
        ws[cell].fill = PatternFill(start_color = 'FFC300', 
                                    end_color = 'FFC300', 
                                    fill_type = "solid")
        
    if exp_hl + e_hl:
            ws.sheet_properties.tabColor = Color(rgb='FFC300')
            wb.save(file)
            
    return 

#-----------------------------------------------------------------------

def rearrange_results (results_df): 
    '''Auxiliar funtion to re organice the output results table so it is 
    comparable with the results given from the excel sheet DP4+ Custom. 
    It is usefull to validate this app
    '''
    new = pd.DataFrame(columns = results_df.columns)
    for row in ['H_sca','C_sca','Sca',
                'H_uns','C_uns','Uns',
                'H_full','C_full','Full']: 
        new.loc [row] = results_df.loc[row]
        
    return new