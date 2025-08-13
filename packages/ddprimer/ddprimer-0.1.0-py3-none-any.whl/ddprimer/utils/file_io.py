#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
File I/O module for ddPrimer pipeline.

Contains functionality for:
1. Cross-platform file selection with GUI and CLI support
2. FASTA file reading and writing operations
3. Excel file formatting with comprehensive styling
4. Sequence table loading from CSV and Excel formats

This module provides consistent interfaces for file operations including
file selection, reading/writing, and format conversions across different
platforms and environments.
"""

import os
import sys
import platform
import logging
import pandas as pd
import contextlib
from typing import Dict, Optional
from pathlib import Path

# Import package modules
from ..config import FileSelectionError, FileFormatError, FileError

# Set up module logger
logger = logging.getLogger(__name__)

# Optional import for Excel formatting
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# Optional import for wxPython file dialogs
try:
    import wx
    HAS_WX = True
except ImportError:
    HAS_WX = False


@contextlib.contextmanager
def _silence_cocoa_stderr():
    """
    Temporarily redirect C-level stderr to /dev/null (macOS IMK chatter).
    
    Yields:
        None - Context manager for silencing stderr on macOS
    """
    if platform.system() != "Darwin":
        yield
        return
        
    import os, sys
    old_fd = os.dup(2)
    devnull = os.open(os.devnull, os.O_WRONLY)
    os.dup2(devnull, 2)
    os.close(devnull)
    try:
        yield
    finally:
        os.dup2(old_fd, 2)
        os.close(old_fd)


class FileIO:
    """
    Handles all file I/O operations with consistent interfaces.
    
    This class provides cross-platform file operations including GUI and CLI
    file selection, FASTA file processing, Excel formatting, and sequence
    table loading. It automatically detects the environment and chooses
    appropriate methods for file operations.
    
    Example:
        >>> sequences = FileIO.load_fasta("genome.fasta")
        >>> file_path = FileIO.select_file("Select genome file", [("FASTA", "*.fasta")])
        >>> FileIO.save_results(df, "output_dir", "input.fasta")
    """
    
    #############################################################################
    #                           Workflow Wrappers
    #############################################################################
    
    @classmethod
    def select_input_files_workflow(cls) -> Dict[str, Optional[str]]:
        """
        Interactive file selection for pipeline input files with workflow integration.
        
        Guides user through selecting required input files (FASTA, VCF, GFF)
        with appropriate file type filtering and validation.
        
        Returns:
            Dictionary with selected file paths: {'fasta': path, 'vcf': path, 'gff': path}
            Values are None if files not selected or selection canceled
            
        Raises:
            FileSelectionError: If file selection fails
        """
        logger.debug("=== WORKFLOW: INPUT FILE SELECTION ===")
        
        try:
            selected_files = {
                'fasta': None,
                'vcf': None, 
                'gff': None
            }
            
            # Select FASTA file
            logger.debug("Requesting FASTA file selection")
            fasta_file = cls.select_file(
                "Select reference genome FASTA file",
                [("FASTA Files", "*.fasta"), ("FASTA Files", "*.fa"), ("FASTA Files", "*.fna")]
            )
            selected_files['fasta'] = fasta_file
            
            # Select VCF file
            logger.debug("Requesting VCF file selection")
            vcf_file = cls.select_file(
                "Select VCF variant file",
                [("VCF Files", "*.vcf"), ("Compressed VCF", "*.vcf.gz")]
            )
            selected_files['vcf'] = vcf_file
            
            # Select GFF file (optional)
            logger.debug("Requesting GFF file selection")
            try:
                gff_file = cls.select_file(
                    "Select GFF annotation file (optional - cancel to skip)",
                    [("GFF Files", "*.gff"), ("GFF3 Files", "*.gff3"), ("Compressed GFF", "*.gff.gz")]
                )
                selected_files['gff'] = gff_file
            except FileSelectionError:
                logger.debug("GFF file selection canceled - proceeding without annotations")
                selected_files['gff'] = None
            
            # Mark selection process complete
            cls.mark_selection_complete()
            
            logger.debug(f"File selection complete: FASTA={bool(selected_files['fasta'])}, "
                        f"VCF={bool(selected_files['vcf'])}, GFF={bool(selected_files['gff'])}")
            logger.debug("=== END WORKFLOW: INPUT FILE SELECTION ===")
            
            return selected_files
            
        except Exception as e:
            error_msg = f"Error in file selection workflow: {str(e)}"
            logger.error(error_msg)
            logger.debug(f"Error details: {str(e)}", exc_info=True)
            logger.debug("=== END WORKFLOW: INPUT FILE SELECTION ===")
            raise FileSelectionError(error_msg) from e
    
    @classmethod
    def load_sequences_workflow(cls, file_path: str) -> Dict[str, str]:
        """
        Load sequences from various file formats for workflow integration.
        
        Automatically detects file format (FASTA, CSV, Excel) and loads
        sequences with appropriate parsing and validation.
        
        Args:
            file_path: Path to sequence file
            
        Returns:
            Dictionary mapping sequence IDs to sequences
            
        Raises:
            FileError: If file cannot be accessed
            FileFormatError: If file format is invalid or parsing fails
        """
        logger.debug("=== WORKFLOW: SEQUENCE LOADING ===")
        logger.debug(f"Loading sequences from: {file_path}")
        
        try:
            # Determine file type and load accordingly
            file_ext = Path(file_path).suffix.lower()
            
            if file_ext in ['.fasta', '.fa', '.fna']:
                logger.debug("Loading as FASTA file")
                sequences = cls.load_fasta(file_path)
            elif file_ext in ['.csv', '.xlsx', '.xls']:
                logger.debug("Loading as sequence table")
                sequences = cls.load_sequences_from_table(file_path)
            else:
                error_msg = f"Unsupported file format: {file_ext}"
                logger.error(error_msg)
                raise FileFormatError(error_msg)
            
            if not sequences:
                error_msg = f"No valid sequences found in file: {file_path}"
                logger.error(error_msg)
                raise FileFormatError(error_msg)
            
            logger.debug(f"Successfully loaded {len(sequences)} sequences")
            logger.debug("=== END WORKFLOW: SEQUENCE LOADING ===")
            
            return sequences
            
        except Exception as e:
            error_msg = f"Error in sequence loading workflow: {str(e)}"
            logger.error(error_msg)
            logger.debug(f"Error details: {str(e)}", exc_info=True)
            logger.debug("=== END WORKFLOW: SEQUENCE LOADING ===")
            raise
    
    @classmethod
    def save_results_workflow(cls, df: pd.DataFrame, output_dir: str, 
                             input_file: str, mode: str = 'standard') -> str:
        """
        Save pipeline results to formatted Excel file for workflow integration.
        
        Creates properly formatted Excel output with comprehensive styling
        and appropriate column organization based on pipeline mode.
        
        Args:
            df: Results DataFrame
            output_dir: Output directory path
            input_file: Original input file path (for naming)
            mode: Pipeline mode ('standard' or 'direct')
            
        Returns:
            Path to saved Excel file
            
        Raises:
            FileFormatError: If file cannot be saved
        """
        logger.debug("=== WORKFLOW: RESULTS SAVING ===")
        logger.debug(f"Saving {len(df)} results to: {output_dir}")
        
        try:
            output_path = cls.save_results(df, output_dir, input_file, mode)
            
            logger.debug(f"Results successfully saved to: {output_path}")
            logger.debug("=== END WORKFLOW: RESULTS SAVING ===")
            
            return output_path
            
        except Exception as e:
            error_msg = f"Error in results saving workflow: {str(e)}"
            logger.error(error_msg)
            logger.debug(f"Error details: {str(e)}", exc_info=True)
            logger.debug("=== END WORKFLOW: RESULTS SAVING ===")
            raise
    
    #############################################################################
    
    # Default to GUI mode unless explicitly detected as headless
    use_cli = False

    # Check for macOS and PyObjC availability
    is_macos = platform.system() == "Darwin"
    has_pyobjc = False
    
    if is_macos:
        try:
            import Foundation
            import AppKit
            has_pyobjc = True
        except ImportError:
            has_pyobjc = False

    # Detect headless environments
    if platform.system() == "Linux":
        if not os.environ.get("DISPLAY", ""):
            use_cli = True
    elif platform.system() == "Windows":
        if not sys.stdout.isatty():
            use_cli = True

    # Initialize last directory
    _last_directory = None
    
    # Shared wxPython app instance
    _wx_app = None
    
    @classmethod
    def initialize_wx_app(cls):
        """
        Initialize the wxPython app if it doesn't exist yet.
        
        Raises:
            FileSelectionError: If wxPython initialization fails
        """
        if HAS_WX and cls._wx_app is None:
            try:
                with _silence_cocoa_stderr():
                    cls._wx_app = wx.App(False)
                    logger.debug("wxPython app initialized successfully")
            except Exception as e:
                error_msg = f"Failed to initialize wxPython app: {str(e)}"
                logger.error(error_msg)
                logger.debug(f"Error details: {str(e)}", exc_info=True)
                raise FileSelectionError(error_msg) from e
    
    @classmethod
    def hide_app(cls):
        """Hide the wxPython app without destroying it."""
        if cls.is_macos and cls.has_pyobjc and cls._wx_app is not None:
            try:
                import AppKit
                NSApplication = AppKit.NSApplication.sharedApplication()
                NSApplication.setActivationPolicy_(AppKit.NSApplicationActivationPolicyAccessory)
                logger.debug("App hidden from macOS dock")
            except Exception as e:
                logger.warning(f"Error hiding app from dock: {str(e)}")
                logger.debug(f"Error details: {str(e)}", exc_info=True)

    @classmethod
    def mark_selection_complete(cls):
        """Mark that all file selections are complete and hide the app."""
        logger.debug("File selection process marked as complete")
        cls.hide_app()

    @classmethod
    def get_last_directory(cls):
        """
        Get the last directory used, loading from unified config storage.
        
        Returns:
            Path to the last directory used, or home directory as fallback
        """
        if cls._last_directory is None:
            try:
                from ..config import Config
                last_dir = Config.load_user_setting("last_directory", None)
                
                if last_dir and os.path.isdir(last_dir):
                    cls._last_directory = last_dir
                    logger.debug(f"Loaded last directory from config: {last_dir}")
                else:
                    cls._last_directory = os.path.expanduser("~")
                    logger.debug(f"Using home directory: {cls._last_directory}")
                    
            except Exception as e:
                logger.warning(f"Error loading last directory: {str(e)}")
                logger.debug(f"Error details: {str(e)}", exc_info=True)
                cls._last_directory = os.path.expanduser("~")
        
        return cls._last_directory
    
    @classmethod
    def save_last_directory(cls, directory):
        """
        Save the last directory to unified config storage.
        
        Args:
            directory: Path to save as last used directory
        """
        if not directory or not isinstance(directory, str) or not os.path.isdir(directory):
            return
            
        cls._last_directory = directory
        
        try:
            from ..config import Config
            Config.save_user_setting("last_directory", directory)
            logger.debug(f"Saved last directory to config: {directory}")
        except Exception as e:
            logger.warning(f"Error saving last directory: {str(e)}")
            logger.debug(f"Error details: {str(e)}", exc_info=True)

    @classmethod
    def normalize_filetypes(cls, filetypes):
        """
        Normalize file types for different platforms.
        
        Args:
            filetypes: List of (description, extension) tuples
            
        Returns:
            Normalized file types list for the current platform
        """
        # Define "All Files" for different platforms
        if cls.is_macos or platform.system() == "Linux":
            all_files = ("All Files", "*")
        else:  # Windows
            all_files = ("All Files", "*.*")
            
        normalized = [all_files]
        
        if filetypes:
            for desc, ext in filetypes:
                if ext == "*" or ext == "*.*":
                    continue
                
                if cls.is_macos:
                    clean_ext = ext.lstrip("*.")
                    if "." in clean_ext:
                        base_ext = clean_ext.split(".")[-1]
                        if base_ext:
                            normalized.insert(0, (desc, base_ext))
                    elif clean_ext:
                        normalized.insert(0, (desc, clean_ext))
                else:
                    if not ext.startswith("*.") and ext != "*":
                        clean_ext = f"*.{ext.lstrip('.')}"
                    else:
                        clean_ext = ext
                    normalized.insert(0, (desc, clean_ext))
                    
        return normalized
        
    @classmethod
    def select_file(cls, prompt, filetypes):
        """
        Show a file dialog or prompt for a file path if CLI mode is enabled.
        
        Args:
            prompt: Text to display in the file dialog
            filetypes: List of file type tuples for the dialog
            
        Returns:
            Selected file path
            
        Raises:
            FileSelectionError: If file selection fails or user cancels in CLI mode
        """
        last_directory = cls.get_last_directory()

        if cls.use_cli:
            try:
                file_path = input(f"{prompt}: ").strip()
                if not file_path:
                    error_msg = "No file path provided in CLI mode"
                    logger.error(error_msg)
                    raise FileSelectionError(error_msg)
                    
                if not os.path.exists(file_path):
                    error_msg = f"File not found: {file_path}"
                    logger.error(error_msg)
                    raise FileSelectionError(error_msg)
                
                return file_path
            except KeyboardInterrupt:
                error_msg = "File selection canceled by user"
                logger.error(error_msg)
                raise FileSelectionError(error_msg)
            except Exception as e:
                error_msg = f"CLI file selection failed: {str(e)}"
                logger.error(error_msg)
                logger.debug(f"Error details: {str(e)}", exc_info=True)
                raise FileSelectionError(error_msg) from e

        # wxPython implementation
        if not cls.use_cli and HAS_WX:
            try:
                cls.initialize_wx_app()
                
                valid_filetypes = cls.normalize_filetypes(filetypes)
                specific_types = [ft for ft in valid_filetypes if not ft[0].startswith("All Files")]
                wildcard = "|".join([f"{desc} ({ext})|{ext}" for desc, ext in specific_types]) or "All Files (*.*)|*.*"

                with _silence_cocoa_stderr():
                    dlg = wx.FileDialog(
                        None,
                        prompt,
                        wildcard=wildcard,
                        style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST
                    )
                    dlg.SetDirectory(last_directory)
                    dlg.SetFilterIndex(0)

                    if dlg.ShowModal() == wx.ID_OK:
                        file_path = dlg.GetPath()
                    else:
                        file_path = ""
                    dlg.Destroy()

                if not file_path:
                    # Dialog was canceled - fall back to CLI instead of raising error
                    logger.warning("File dialog was canceled, falling back to CLI mode")
                    cls.use_cli = True
                    return cls.select_file(prompt, filetypes)  # Recursive call in CLI mode

                cls.save_last_directory(os.path.dirname(file_path))
                return file_path

            except FileSelectionError:
                raise
            except Exception as e:
                logger.warning(f"wxPython file selection failed: {str(e)}")
                logger.debug(f"Error details: {str(e)}", exc_info=True)
                cls.use_cli = True
        
        # Fallback to CLI
        logger.warning("Falling back to CLI mode for file selection")
        cls.use_cli = True
        return cls.select_file(prompt, filetypes)
        
    @staticmethod
    def load_fasta(filepath):
        """
        Load sequences from a FASTA file into a dictionary.
        
        Args:
            filepath: Path to the FASTA file
            
        Returns:
            Dictionary mapping sequence headers to sequences
            
        Raises:
            FileError: If the FASTA file doesn't exist
            FileFormatError: If there's an error parsing the FASTA file
        """
        sequences = {}
        name = None
        seq_chunks = []
        
        if not os.path.exists(filepath):
            error_msg = f"FASTA file not found: {filepath}"
            logger.error(error_msg)
            raise FileError(error_msg)
            
        try:
            with open(filepath, 'r') as f:
                for line in f:
                    line = line.strip()
                    if not line:
                        continue
                    if line.startswith(">"):
                        if name:
                            sequences[name] = "".join(seq_chunks).upper()
                        name = line[1:].split()[0]
                        seq_chunks = []
                    else:
                        seq_chunks.append(line)
                        
                # Handle the last sequence
                if name:
                    sequences[name] = "".join(seq_chunks).upper()
                    
        except (OSError, IOError) as e:
            error_msg = f"Error reading FASTA file {os.path.abspath(filepath)}: {str(e)}"
            logger.error(error_msg)
            logger.debug(f"Error details: {str(e)}", exc_info=True)
            raise FileError(error_msg) from e
        except Exception as e:
            error_msg = f"Error parsing FASTA file {os.path.abspath(filepath)}: {str(e)}"
            logger.error(error_msg)
            logger.debug(f"Error details: {str(e)}", exc_info=True)
            raise FileFormatError(error_msg) from e
        
        logger.debug(f"Successfully loaded {len(sequences)} sequences from FASTA file")
        return sequences
    
    @staticmethod
    def load_sequences_from_table(file_path):
        """
        Load sequences from CSV or Excel file with flexible column detection.
        
        This method provides compatibility with the existing FileIO interface
        while using the enhanced DirectModeProcessor for flexible column detection.
        
        Args:
            file_path: Path to CSV or Excel file containing sequences
            
        Returns:
            Dictionary mapping sequence IDs to sequences
            
        Raises:
            FileError: If file cannot be accessed
            FileFormatError: If file format is invalid or columns cannot be detected
            SequenceProcessingError: If sequences are invalid
        """
        try:
            # Import DirectModeProcessor to avoid circular imports
            from ..utils.direct_mode import DirectModeProcessor
            
            # Delegate to the DirectModeProcessor for enhanced functionality
            return DirectModeProcessor.load_sequences_from_table(file_path)
        except ImportError:
            # Fallback for cases where direct_mode is not available
            error_msg = f"Direct mode functionality not available. Cannot load sequences from table: {file_path}"
            logger.error(error_msg)
            raise FileFormatError(error_msg)
    
    def _prepare_output_dataframe(df, mode='standard'):
        """
        Prepare DataFrame for output by formatting columns and adding derived fields.
        
        Args:
            df: Input DataFrame with primer records
            mode: Pipeline mode ('standard' or 'direct')
            
        Returns:
            DataFrame with properly formatted columns for output
        """
        output_df = df.copy()
        
        # 1. AMPLICON LENGTH: Create 'Length' from 'Amplicon Length' if it doesn't exist
        if 'Length' not in output_df.columns:
            if 'Amplicon Length' in output_df.columns:
                output_df['Length'] = output_df['Amplicon Length']
            elif 'Sequence (A)' in output_df.columns:
                output_df['Length'] = output_df['Sequence (A)'].apply(lambda x: len(str(x)) if pd.notna(x) else 0)

        # 2. Calculate GC% for amplicons using existing FilterProcessor method
        if 'GC%' not in output_df.columns and 'Sequence (A)' in output_df.columns:
            from ..core.filter_processor import FilterProcessor
            logger.debug("Calculating GC% for amplicon sequences")
            output_df['GC%'] = output_df['Sequence (A)'].apply(FilterProcessor.calculate_gc)

        # 3. LOCATION: Use only 'Start' coordinate
        if 'Start' in output_df.columns:
            output_df['Location'] = output_df['Start'].apply(
                lambda start: str(int(start)) if pd.notna(start) else ""
            )
        
        return output_df

    @staticmethod
    def format_excel(df, output_file):
        """
        Save DataFrame to Excel with comprehensive formatting including number formats.
        
        Args:
            df: DataFrame with primer results
            output_file: Path to save the formatted Excel file
            
        Returns:
            Path to the saved Excel file
            
        Raises:
            FileFormatError: If Excel file cannot be created
        """
        try:
            if not HAS_OPENPYXL:
                raise ImportError("openpyxl is not available")
                
            # First, save with pandas to get a basic Excel file
            df.to_excel(output_file, index=False, engine='openpyxl')
            
            # Now open the file for formatting
            workbook = openpyxl.load_workbook(output_file)
            worksheet = workbook.active
            
            max_row = worksheet.max_row
            max_col = worksheet.max_column
            
            # Create a new row for our custom headers
            worksheet.insert_rows(1)
            max_row += 1
            
            # Create styles
            header_font = Font(bold=True)
            sequence_fill = PatternFill(
                start_color='D9D9D9',
                end_color='D9D9D9',
                fill_type='solid'
            )

            centered_alignment = Alignment(horizontal='center', vertical='center')
            left_alignment = Alignment(horizontal='left', vertical='center')
            
            # Create border styles for group separators
            thin_border_left = Border(left=Side(style='thin'))
            thin_border_right = Border(right=Side(style='thin'))
            thin_border_both = Border(left=Side(style='thin'), right=Side(style='thin'))
            
            # Clear all default borders first
            no_border = Border()
            for row_num in range(1, max_row + 1):
                for col_num in range(1, max_col + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.border = no_border
            
            column_map = {}
            header_texts = []
            
            # Apply basic formatting
            for col_num in range(1, max_col + 1):
                col_letter = get_column_letter(col_num)
                
                # Set header row formatting
                cell2 = worksheet.cell(row=2, column=col_num)
                cell2.font = header_font
                cell2.alignment = centered_alignment
                
                # Handle Gene column
                if cell2.value == "Gene":
                    cell1 = worksheet.cell(row=1, column=col_num)
                    cell1.value = "Gene"
                    cell1.font = header_font
                    cell1.alignment = centered_alignment
                    worksheet.merge_cells(start_row=1, start_column=col_num, end_row=2, end_column=col_num)
                    # Set wider width for Gene column
                    worksheet.column_dimensions[col_letter].width = 15
                else:
                    # Set column widths based on content type
                    header_text = cell2.value
                    if header_text in ["Sequence (F)", "Sequence (R)", "Sequence (P)", "Sequence (A)"]:
                        worksheet.column_dimensions[col_letter].width = 15
                    elif header_text == "Match_Quality":
                        worksheet.column_dimensions[col_letter].width = 12
                    else:
                        worksheet.column_dimensions[col_letter].width = 10

                header_text = cell2.value
                header_texts.append(header_text)
                column_map[header_text] = col_num
                
                # Apply number formatting and special styling based on column type
                for row_num in range(3, max_row + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    
                    # Apply sequence fill to sequence columns
                    if header_text in ["Sequence (F)", "Sequence (R)", "Sequence (P)", "Sequence (A)"]:
                        cell.fill = sequence_fill
                    
                    # Apply number formatting
                    elif header_text and "BLAST" in header_text:
                        # Scientific notation for BLAST columns (e.g., 1.23E-05)
                        cell.number_format = '0.00E+00'
                        cell.alignment = centered_alignment
                    elif header_text and any(col_type in header_text for col_type in ["Tm (", "Penalty (", "dG ("]):
                        # One decimal place for Tm, Penalty, and dG columns
                        cell.number_format = '0.0'
                        cell.alignment = centered_alignment
                    elif header_text == "Length":
                        # Integer format for Length
                        cell.number_format = '0'
                        cell.alignment = centered_alignment
                    elif header_text == "GC%":
                        # One decimal place for GC%
                        cell.number_format = '0.0'
                        cell.alignment = centered_alignment
                    else:
                        # Handle "No suitable primers found" cells and other text
                        if cell.value == "No suitable primers found":
                            cell.alignment = left_alignment
                        else:
                            cell.alignment = centered_alignment
            
            # Freeze panes
            worksheet.freeze_panes = 'B3'
            
            # Group columns with updated header names using position indicators
            header_groups = {
                "Gene": [],
                "Forward Primer": ["Sequence (F)", "Tm (F)", "Penalty (F)", "dG (F)", "BLAST (F)"],
                "Reverse Primer": ["Sequence (R)", "Tm (R)", "Penalty (R)", "dG (R)", "BLAST (R)"],
                "Probe": ["Sequence (P)", "Tm (P)", "Penalty (P)", "dG (P)", "BLAST (P)"],
                "Amplicon": ["Sequence (A)", "Length", "GC%", "dG (A)"],
                "Location": ["Chr", "Location", "Match_Quality"]
            }
            
            # Track group boundaries for border application
            group_boundaries = []
            # Track which columns have been assigned to groups to avoid duplicates
            assigned_columns = set()
            
            for group_name, headers in header_groups.items():
                # Only include headers that exist in the DataFrame and haven't been assigned yet
                existing_headers = [h for h in headers if h in header_texts and h not in assigned_columns]
                
                if not existing_headers:
                    continue
                    
                # Mark these columns as assigned
                assigned_columns.update(existing_headers)
                
                col_indices = [column_map[h] for h in existing_headers]
                
                if col_indices:
                    start_col = min(col_indices)
                    end_col = max(col_indices)
                    
                    # Store group boundaries
                    group_boundaries.append((start_col, end_col))
                    
                    group_cell = worksheet.cell(row=1, column=start_col)
                    group_cell.value = group_name
                    group_cell.font = header_font
                    group_cell.alignment = centered_alignment
                    
                    if start_col != end_col:
                        merge_range = f"{get_column_letter(start_col)}1:{get_column_letter(end_col)}1"
                        try:
                            worksheet.merge_cells(merge_range)
                        except Exception as e:
                            logger.warning(f"Could not merge range {merge_range}: {str(e)}")
            
            # Apply borders to group boundaries
            for start_col, end_col in group_boundaries:
                # Apply left border to start of group
                for row_num in range(1, max_row + 1):
                    cell = worksheet.cell(row=row_num, column=start_col)
                    cell.border = Border(
                        left=Side(style='thin'),
                        top=cell.border.top,
                        bottom=cell.border.bottom,
                        right=cell.border.right
                    )
                
                # Apply right border to end of group
                for row_num in range(1, max_row + 1):
                    cell = worksheet.cell(row=row_num, column=end_col)
                    cell.border = Border(
                        left=cell.border.left,
                        right=Side(style='thin'),
                        top=cell.border.top,
                        bottom=cell.border.bottom
                    )
            
            # Add medium border around the entire populated table
            medium_border_side = Side(style='medium')
            
            # Top border
            for col_num in range(1, max_col + 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.border = Border(
                    top=medium_border_side,
                    left=cell.border.left,
                    right=cell.border.right,
                    bottom=cell.border.bottom
                )
            
            # Bottom border
            for col_num in range(1, max_col + 1):
                cell = worksheet.cell(row=max_row, column=col_num)
                cell.border = Border(
                    bottom=medium_border_side,
                    left=cell.border.left,
                    right=cell.border.right,
                    top=cell.border.top
                )
            
            # Left border
            for row_num in range(1, max_row + 1):
                cell = worksheet.cell(row=row_num, column=1)
                cell.border = Border(
                    left=medium_border_side,
                    top=cell.border.top,
                    right=cell.border.right,
                    bottom=cell.border.bottom
                )
            
            # Right border
            for row_num in range(1, max_row + 1):
                cell = worksheet.cell(row=row_num, column=max_col)
                cell.border = Border(
                    right=medium_border_side,
                    left=cell.border.left,
                    top=cell.border.top,
                    bottom=cell.border.bottom
                )
            
            workbook.save(output_file)
            return output_file
                
        except ImportError:
            logger.warning("openpyxl not available, falling back to standard Excel export")
        except Exception as e:
            error_msg = f"Error applying Excel formatting: {str(e)}"
            logger.error(error_msg)
            logger.warning("Falling back to standard Excel export")
            logger.debug(f"Error details: {str(e)}", exc_info=True)
            
        # Basic fallback save
        try:
            df.to_excel(output_file, index=False)
            logger.info(f"Excel file saved to: {output_file} (without formatting)")
            return output_file
        except Exception as e:
            error_msg = f"Failed to save Excel file {output_file}: {str(e)}"
            logger.error(error_msg)
            logger.debug(f"Error details: {str(e)}", exc_info=True)
            raise FileFormatError(error_msg) from e

    @staticmethod
    def save_results(df, output_dir, input_file, mode='standard', chromosome_map=None):
        """
        Save results to an Excel file with correct naming and standardized chromosome output.

        Args:
            df: DataFrame with primer results.
            output_dir: Output directory.
            input_file: Path to the input file (FASTA, CSV, etc.).
            mode: Pipeline mode ('standard' or 'direct').
            chromosome_map (dict, optional): Map of original to standardized chromosome names.

        Returns:
            Path to the output file.

        Raises:
            FileFormatError: If there's an error saving the results.
        """
        # Make sure output directory exists
        try:
            os.makedirs(output_dir, exist_ok=True)
        except OSError as e:
            error_msg = f"Failed to create output directory {output_dir}: {str(e)}"
            logger.error(error_msg)
            logger.debug(f"Error details: {str(e)}", exc_info=True)
            raise FileFormatError(error_msg) from e

        # Apply chromosome name standardization for the output file
        if chromosome_map and 'Chr' in df.columns:
            logger.debug("Applying standardized chromosome names to output.")
            df['Chr'] = df['Chr'].map(chromosome_map).fillna(df['Chr'])

        # Set output filename based on mode
        basename = os.path.basename(input_file)
        root, _ = os.path.splitext(basename)
        
        if mode == 'remap':
            output_filename = f"{root}_remapped.xlsx"
        else:
            output_filename = f"Primers_{root}.xlsx"
            
        output_file = os.path.join(output_dir, output_filename)
        
        # Prepare dataframe
        df = FileIO._prepare_output_dataframe(df, mode)

        # Define final column order with new column names - include probe columns from the start
        base_columns = [
            "Gene",
            "Sequence (F)", "Tm (F)", "Penalty (F)", "dG (F)", "BLAST (F)",
            "Sequence (R)", "Tm (R)", "Penalty (R)", "dG (R)", "BLAST (R)"
        ]
        
        # Add probe columns if present
        if "Sequence (P)" in df.columns:
            probe_cols = ["Sequence (P)", "Tm (P)", "Penalty (P)", "dG (P)", "BLAST (P)"]
            base_columns.extend(probe_cols)
        
        # Add amplicon columns
        amplicon_cols = ["Sequence (A)", "Length", "GC%", "dG (A)"]
        base_columns.extend(amplicon_cols)
        
        # Define final column order based on mode
        if mode == 'direct':
            expected_columns = base_columns
        elif mode == 'remap':
            # More flexible column ordering for remap mode - include all available columns
            expected_columns = base_columns + ["Chr", "Location", "Match_Quality"]
            
            # Add any additional columns that exist in the DataFrame but aren't in expected_columns
            additional_columns = [col for col in df.columns 
                                if col not in expected_columns 
                                and col != "Start"]  # Explicitly exclude "Start" column
            if additional_columns:
                logger.debug(f"Including additional columns in remap output: {additional_columns}")
                expected_columns.extend(additional_columns)
                
        else:  # standard mode
            expected_columns = base_columns + ["Chr", "Location"]
        
        # Only select columns that actually exist in the DataFrame
        available_columns = [col for col in expected_columns if col in df.columns]
        missing_columns = [col for col in expected_columns if col not in df.columns]
        
        if missing_columns and mode != 'remap':
            # Only warn about missing columns in non-remap modes
            logger.warning(f"Missing expected columns: {missing_columns}")
            logger.debug(f"Available columns: {df.columns.tolist()}")
        elif mode == 'remap':
            # In remap mode, just inform about what's available
            logger.debug(f"Remap mode: Using {len(available_columns)} available columns")
            logger.debug(f"Available columns: {available_columns}")
        
        logger.debug(f"Final column selection: {available_columns}")
        df = df[available_columns]
        
        # Save with formatting
        try:
            output_path = FileIO.format_excel(df, output_file)
            return output_path
        except Exception as e:
            error_msg = f"Error saving Excel file: {str(e)}"
            logger.error(error_msg)
            logger.warning("Falling back to basic Excel export")
            logger.debug(f"Error details: {str(e)}", exc_info=True)
            
            try:
                df.to_excel(output_file, index=False)
                logger.info(f"\nResults saved to: {output_file} (without formatting)")
                return output_file
            except Exception as ex:
                error_msg = f"Failed to save results to {output_file}: {str(ex)}"
                logger.error(error_msg)
                logger.debug(f"Error details: {str(ex)}", exc_info=True)
                raise FileFormatError(error_msg) from ex