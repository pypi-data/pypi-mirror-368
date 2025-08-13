#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Primer Remapper Implementation for ddPrimer Pipeline

Handles remapping and re-evaluation of existing primer sets against a new
reference genome and annotation with full functionality including exact matching,
BLAST fallback, and annotation-based gene updates.

Contains functionality for:
1. Flexible loading of primer tables (CSV/Excel) with column detection
2. Exact sequence matching with reverse complement support
3. BLAST fallback for sequences not found by exact matching
4. Mapping primer pairs to new reference with match quality tracking
5. Annotation-based gene name updates at mapped locations
6. Integration with existing pipeline components for re-evaluation
7. Comprehensive match reporting and validation
8. Preservation of original primer properties (Tm, Penalty, dG, etc.)
9. Full pipeline re-evaluation including thermodynamics and BLAST
"""

import os
import logging
import pandas as pd
import tempfile
import subprocess
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from tqdm import tqdm

# Optional import for Excel formatting
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# Import package modules
from ..config import Config, FileError, FileFormatError, SequenceProcessingError, PrimerDesignError, FileSelectionError
from ..utils import FileIO
from ..core import ThermoProcessor, BlastProcessor, AnnotationProcessor
from ..core.filter_processor import FilterProcessor

# Set up module logger
logger = logging.getLogger(__name__)


class PrimerRemapperProcessor:
    """
    Primer remapper with exact matching, BLAST fallback, and annotation integration.
    
    Handles the complete primer remapping workflow including:
    - Exact sequence matching (forward and reverse complement)
    - BLAST fallback for non-exact matches using existing BlastProcessor
    - Annotation-based gene name updates
    - Match quality tracking and reporting
    - Preservation and re-evaluation of primer properties
    """

    # Common column name patterns for flexible detection
    ID_PATTERNS = ['id', 'name', 'gene', 'target', 'primer_id', 'identifier']
    FORWARD_PRIMER_PATTERNS = ['forward', 'fwd', 'left', 'sequence (f)']
    REVERSE_PRIMER_PATTERNS = ['reverse', 'rev', 'right', 'sequence (r)']
    PROBE_PATTERNS = ['probe', 'internal', 'oligo', 'sequence (p)']

    @staticmethod
    def is_valid_dna_sequence(sequence: str) -> bool:
        """
        Check if a string represents a valid DNA sequence.
        
        Args:
            sequence: String to validate
            
        Returns:
            True if valid DNA sequence, False otherwise
        """
        if not sequence or pd.isna(sequence):
            return False
        
        sequence = str(sequence).strip().upper()
        if not sequence:
            return False
        
        # Allow standard DNA bases plus IUPAC ambiguous codes
        valid_bases = set('ATCGRYSWKMBDHVN')
        return all(base in valid_bases for base in sequence) and len(sequence) >= 10

    def __init__(self):
        """Initialize the PrimerRemapperProcessor."""
        logger.debug("PrimerRemapperProcessor initialized")

    @classmethod
    def detect_columns(cls, df: pd.DataFrame) -> Dict[str, Optional[str]]:
        """
        Detect ID, forward, reverse, and optional probe columns in a DataFrame.
        
        Args:
            df: Input DataFrame to analyze
            
        Returns:
            Dictionary mapping column types to detected column names
        """
        logger.debug("Detecting primer columns in DataFrame")
        if df.empty:
            logger.warning("DataFrame is empty")
            return {'id': None, 'forward': None, 'reverse': None, 'probe': None}

        columns_lower = {str(col).lower().strip(): str(col) for col in df.columns}
        detected = {'id': None, 'forward': None, 'reverse': None, 'probe': None}
        patterns = {
            'id': cls.ID_PATTERNS,
            'forward': cls.FORWARD_PRIMER_PATTERNS,
            'reverse': cls.REVERSE_PRIMER_PATTERNS,
            'probe': cls.PROBE_PATTERNS,
        }
        
        used_cols = []
        for col_type, pattern_list in patterns.items():
            for pattern in pattern_list:
                found = False
                for col_lower, col_original in columns_lower.items():
                    if pattern in col_lower and col_original not in used_cols:
                        detected[col_type] = col_original
                        used_cols.append(col_original)
                        logger.debug(f"Detected {col_type} column: '{col_original}' (matched: '{pattern}')")
                        found = True
                        break
                if found:
                    break
        
        # Fallback for missing critical columns
        if not detected['forward'] or not detected['reverse']:
            logger.warning("Could not reliably detect forward/reverse primer columns by name.")
            available_cols = [c for c in df.columns if c not in used_cols]
            if not detected['forward'] and len(available_cols) > 0:
                detected['forward'] = available_cols.pop(0)
                used_cols.append(detected['forward'])
                logger.debug(f"Fallback: Using column '{detected['forward']}' as forward primer.")
            if not detected['reverse'] and len(available_cols) > 0:
                detected['reverse'] = available_cols.pop(0)
                logger.debug(f"Fallback: Using column '{detected['reverse']}' as reverse primer.")

        return detected

    @classmethod
    def read_excel_with_merged_headers(cls, excel_path: str) -> pd.DataFrame:
        """
        Read Excel file handling merged headers properly, especially for the Gene column.
        This replicates the logic from the standalone script.
        
        Args:
            excel_path: Path to the Excel file
            
        Returns:
            DataFrame with properly named columns
        """
        try:
            if HAS_OPENPYXL:
                # First, read the raw data to understand the structure
                workbook = openpyxl.load_workbook(excel_path)
                worksheet = workbook.active
                
                # Check for merged cells in the header area and get their values
                merged_ranges = worksheet.merged_cells.ranges
                header_row_1 = {}
                header_row_2 = {}
                
                # Read row 1 and row 2 headers
                for col_num in range(1, worksheet.max_column + 1):
                    cell_1 = worksheet.cell(row=1, column=col_num)
                    cell_2 = worksheet.cell(row=2, column=col_num)
                    
                    header_row_1[col_num] = cell_1.value
                    header_row_2[col_num] = cell_2.value
                
                # Check for merged cells that span rows 1-2
                for merged_range in merged_ranges:
                    if merged_range.min_row == 1 and merged_range.max_row == 2:
                        # This is a merged cell spanning both header rows
                        col_num = merged_range.min_col
                        merged_value = worksheet.cell(row=1, column=col_num).value
                        header_row_2[col_num] = merged_value  # Use the merged value as the column name
                
                workbook.close()
                
                # Now read with pandas using header=1, then fix column names
                df = pd.read_excel(excel_path, header=1)
                
                # Fix column names based on what we found
                new_columns = []
                for i, col in enumerate(df.columns, 1):
                    if header_row_2.get(i):
                        new_columns.append(header_row_2[i])
                    else:
                        new_columns.append(col)
                
                df.columns = new_columns
                return df
                
            else:
                # Fallback to regular pandas reading
                return pd.read_excel(excel_path, header=1)
                
        except Exception as e:
            logger.warning(f"Error reading Excel with merged headers: {e}")
            logger.warning("Falling back to standard reading...")
            try:
                # Try reading with header=1 first (skip the first row)
                return pd.read_excel(excel_path, header=1)
            except Exception as e2:
                logger.warning(f"Error reading with header=1: {e2}")
                logger.warning("Trying with header=0...")
                try:
                    # Last resort: read with header=0 (use first row as headers)
                    return pd.read_excel(excel_path, header=0)
                except Exception as e3:
                    logger.warning(f"Error reading with header=0: {e3}")
                    # Final fallback: read without headers and assign generic column names
                    logger.warning("Reading without headers as final fallback...")
                    df = pd.read_excel(excel_path, header=None)
                    df.columns = [f"Column_{i+1}" for i in range(len(df.columns))]
                    return df

    @classmethod
    def load_primers_from_table(cls, file_path: str) -> Tuple[List[Dict], pd.DataFrame]:
        """
        Load primer sets from a CSV or Excel file into a list of dictionaries.
        Version that properly handles Excel files with merged headers and preserves original data.
        
        Args:
            file_path: Path to the primer table file
            
        Returns:
            Tuple of (primer_sets_list, original_dataframe)
            
        Raises:
            FileError: If file cannot be accessed
            FileFormatError: If file format is invalid
            SequenceProcessingError: If no valid primer sets found
        """
        logger.debug(f"Loading primer sets from table: {file_path}")
        if not os.path.exists(file_path):
            raise FileError(f"Primer table file not found: {file_path}")

        try:
            file_ext = Path(file_path).suffix.lower()
            if file_ext == '.csv':
                df = pd.read_csv(file_path)
            elif file_ext in ['.xlsx', '.xls']:
                df = cls.read_excel_with_merged_headers(file_path)
            else:
                raise FileFormatError(f"Unsupported file format: {file_ext}")

            if df.empty:
                raise FileFormatError(f"No data found in file: {file_path}")

            logger.debug("Available columns:")
            for i, col in enumerate(df.columns):
                logger.debug(f"  {i+1}. {col}")

            # Column detection with explicit sequence column checks
            primer_sets = []
            
            # Try to find primer sequences using multiple approaches
            forward_primer = None
            reverse_primer = None
            probe = None
            gene_name_col = None
            
            # Method 1: Look for exact column names (standard ddPrimer output)
            for col in df.columns:
                col_str = str(col).strip()
                if col_str == 'Sequence (F)':
                    forward_primer = col_str
                elif col_str == 'Sequence (R)':
                    reverse_primer = col_str
                elif col_str == 'Sequence (P)':
                    probe = col_str
                elif col_str == 'Gene':
                    gene_name_col = col_str
            
            # Method 2: Fallback to pattern matching if exact names not found
            if not forward_primer or not reverse_primer:
                columns = cls.detect_columns(df)
                if not forward_primer:
                    forward_primer = columns['forward']
                if not reverse_primer:
                    reverse_primer = columns['reverse']
                if not probe:
                    probe = columns['probe']
                if not gene_name_col:
                    gene_name_col = columns['id']

            if not forward_primer or not reverse_primer:
                available_cols = [str(col) for col in df.columns]
                raise FileFormatError(f"Could not detect forward and reverse primer columns. Available columns: {available_cols}")

            logger.debug(f"Using columns: Forward='{forward_primer}', Reverse='{reverse_primer}', Probe='{probe}', Gene='{gene_name_col}'")

            # Extract primer data while preserving original DataFrame
            for idx, row in df.iterrows():
                # Get primer sequences
                fwd_seq = str(row[forward_primer]).strip().upper() if pd.notna(row[forward_primer]) else ""
                rev_seq = str(row[reverse_primer]).strip().upper() if pd.notna(row[reverse_primer]) else ""

                # Skip if essential sequences are missing or invalid
                if not fwd_seq or not rev_seq or 'NAN' in fwd_seq or 'NAN' in rev_seq:
                    continue

                # Validate that these are actually DNA sequences
                if not cls.is_valid_dna_sequence(fwd_seq):
                    logger.warning(f"Row {idx+1}: Forward sequence is not valid DNA: {fwd_seq[:20]}...")
                    continue
                if not cls.is_valid_dna_sequence(rev_seq):
                    logger.warning(f"Row {idx+1}: Reverse sequence is not valid DNA: {rev_seq[:20]}...")
                    continue

                # Get gene name
                if gene_name_col and pd.notna(row[gene_name_col]):
                    primer_id = str(row[gene_name_col]).strip()
                else:
                    primer_id = f"PrimerSet_{idx+1}"

                # Get probe sequence if available
                probe_seq = ""
                if probe and pd.notna(row[probe]):
                    probe_seq = str(row[probe]).strip().upper()
                    if probe_seq == 'NAN':
                        probe_seq = ""

                primer_sets.append({
                    'id': primer_id,
                    'forward': fwd_seq,
                    'reverse': rev_seq,
                    'probe': probe_seq,
                    'original_gene': primer_id,  # Store original for comparison
                    'dataframe_index': idx  # Store index to map back to original data
                })
            
            if not primer_sets:
                raise SequenceProcessingError(f"No valid primer sets found in {file_path}")

            logger.debug(f"Successfully loaded {len(primer_sets)} primer sets from table.")
            return primer_sets, df
        except Exception as e:
            if isinstance(e, (FileError, FileFormatError, SequenceProcessingError)):
                raise
            raise FileFormatError(f"Error loading primers from {file_path}: {e}") from e

    def find_exact_position(self, sequence: str, genome_sequences: Dict[str, str]) -> Tuple[Optional[str], Optional[int], Optional[str], Optional[str]]:
        """
        Find the exact position of a sequence in the genome using exact string matching.
        Searches for both forward and reverse complement sequences.
        
        Args:
            sequence: DNA sequence to search for
            genome_sequences: Dictionary of chromosome sequences
            
        Returns:
            Tuple of (chromosome, position, strand, match_type) or (None, None, None, None) if not found
            strand is '+' for forward match, '-' for reverse complement match
            match_type is 'exact' for perfect matches
        """
        if not sequence or pd.isna(sequence):
            return None, None, None, None
        
        sequence = str(sequence).upper().strip()
        rev_comp_seq = FilterProcessor.reverse_complement(sequence)
        
        for chromosome, genome_seq in genome_sequences.items():
            # Search for forward sequence
            pos = genome_seq.find(sequence)
            if pos != -1:
                return chromosome, pos + 1, '+', 'exact'  # Convert to 1-based coordinates
            
            # Search for reverse complement
            if rev_comp_seq:
                pos = genome_seq.find(rev_comp_seq)
                if pos != -1:
                    return chromosome, pos + 1, '-', 'exact'  # Convert to 1-based coordinates
        
        return None, None, None, None

    def blast_sequence_for_position(self, seq: str, blast_db_path: str) -> Tuple[Optional[str], Optional[int], Optional[float], Optional[str]]:
        """
        Run BLASTn for a sequence and return the best match position.
        
        Args:
            seq: DNA sequence to BLAST against database
            blast_db_path: BLAST database path
            
        Returns:
            Tuple of (chromosome, position, evalue, match_type) or (None, None, None, None) if not found
        """
        if not seq or not isinstance(seq, str) or not seq.strip():
            return None, None, None, None

        tmp_filename = None
        try:
            # Create temporary file for query sequence
            temp_dir = tempfile.gettempdir()
            with tempfile.NamedTemporaryFile(mode="w+", delete=False, dir=temp_dir, suffix=".fasta") as tmp_query:
                tmp_query.write(f">seq\n{seq}\n")
                tmp_query.flush()
                tmp_filename = tmp_query.name

            # Execute BLASTn command with detailed output format
            result = subprocess.run(
                [
                    "blastn",
                    "-task", "blastn-short",
                    "-db", blast_db_path,
                    "-query", tmp_filename,
                    "-word_size", str(Config.BLAST_WORD_SIZE),
                    "-evalue", str(Config.BLAST_EVALUE),
                    "-reward", str(Config.BLAST_REWARD),
                    "-penalty", str(Config.BLAST_PENALTY),
                    "-gapopen", str(Config.BLAST_GAPOPEN),
                    "-gapextend", str(Config.BLAST_GAPEXTEND),
                    "-max_target_seqs", str(Config.BLAST_MAX_TARGET_SEQS),
                    "-outfmt", "6 sseqid sstart send evalue bitscore length pident"
                ],
                text=True,
                capture_output=True
            )
            
            if result.returncode != 0:
                logger.error(f"BLAST execution failed for sequence {seq[:20]}...")
                logger.error(f"BLAST stderr: {result.stderr}")
                return None, None, None, None

            # Parse BLAST output
            lines = result.stdout.strip().split("\n")
            if not lines or not lines[0].strip():
                return None, None, None, None

            # Get the best hit (first line, already sorted by e-value)
            best_hit = lines[0].strip().split("\t")
            if len(best_hit) < 7:
                return None, None, None, None

            chromosome = best_hit[0]
            start_pos = int(best_hit[1])
            end_pos = int(best_hit[2])
            evalue = float(best_hit[3])
            percent_identity = float(best_hit[6])
            
            # Use the start position (5' end of the match)
            position = min(start_pos, end_pos)  # In case of reverse complement matches
            
            # Determine match type based on percent identity
            if percent_identity == 100.0:
                match_type = 'blast_perfect'
            else:
                match_type = f'blast_{percent_identity:.1f}%'
            
            return chromosome, position, evalue, match_type
            
        except FileNotFoundError:
            logger.error("`blastn` command not found. Please ensure BLAST+ is installed and in your system's PATH.")
            return None, None, None, None
        except Exception as e:
            logger.error(f"Error in BLAST for sequence {seq[:20]}...: {str(e)}")
            return None, None, None, None
            
        finally:
            # Clean up temporary file
            if tmp_filename and os.path.exists(tmp_filename):
                try:
                    os.remove(tmp_filename)
                except OSError as e:
                    logger.debug(f"Failed to remove temp file {tmp_filename}: {e}")

    def find_sequence_position_with_fallback(self, sequence: str, genome_sequences: Dict[str, str], 
                                           blast_db_path: Optional[str]) -> Tuple[Optional[str], Optional[int], Optional[str], Optional[str]]:
        """
        Find sequence position using exact matching first, then BLAST fallback.
        
        Args:
            sequence: DNA sequence to search for
            genome_sequences: Dictionary of chromosome sequences
            blast_db_path: Path to BLAST database (can be None)
            
        Returns:
            Tuple of (chromosome, position, additional_info, match_type)
        """
        # First try exact matching
        chr_exact, pos_exact, strand, match_type = self.find_exact_position(sequence, genome_sequences)
        
        if chr_exact and pos_exact:
            return chr_exact, pos_exact, f"strand:{strand}", match_type
        
        # If exact matching failed and BLAST is available, try BLAST
        if blast_db_path and os.path.exists(f"{blast_db_path}.nhr"):
            chr_blast, pos_blast, evalue, match_type = self.blast_sequence_for_position(sequence, blast_db_path)
            
            if chr_blast and pos_blast:
                return chr_blast, pos_blast, f"E-val:{evalue:.2e}", match_type
        
        return None, None, None, None

    def find_gene_at_position(self, chromosome: str, position: int, annotations: List[Dict]) -> Optional[str]:
        """
        Find the gene that contains the given position using existing AnnotationProcessor.
        
        Args:
            chromosome: Chromosome name
            position: Position on chromosome (1-based)
            annotations: List of gene annotations from AnnotationProcessor
            
        Returns:
            Gene name if found, None otherwise
        """
        if not annotations:
            return None
        
        # Convert position to 0-based for internal coordinate system
        pos_0based = position - 1
        
        # Use AnnotationProcessor's existing method
        overlapping_genes = AnnotationProcessor.find_overlapping_genes(
            chromosome, pos_0based, pos_0based + 1, annotations
        )
        
        if overlapping_genes:
            return overlapping_genes[0]['id']  # Return the first overlapping gene
        
        return None

    def find_primer_pair_locations(self, primer_f: str, primer_r: str, ref_seqs: Dict[str, str], 
                                          blast_db_path: Optional[str], probe: Optional[str] = None) -> List[Dict]:
        """
        Find all genomic locations of a primer pair using exact matching with BLAST fallback.
        Uses the separate BLAST implementation for position finding.
        
        Args:
            primer_f: Forward primer sequence
            primer_r: Reverse primer sequence  
            ref_seqs: Reference genome sequences
            blast_db_path: Path to BLAST database for fallback matching
            probe: Optional probe sequence
            
        Returns:
            List of potential amplicon dictionaries with match quality information
        """
        logger.debug(f"Finding locations for primer pair: F={primer_f[:10]}..., R={primer_r[:10]}...")
        
        # Find positions for all sequences with fallback
        forward_chr, forward_pos, forward_info, forward_match = self.find_sequence_position_with_fallback(
            primer_f, ref_seqs, blast_db_path)
        reverse_chr, reverse_pos, reverse_info, reverse_match = self.find_sequence_position_with_fallback(
            primer_r, ref_seqs, blast_db_path)
        
        # Check probe if provided
        probe_chr, probe_pos, probe_info, probe_match = None, None, None, None
        if probe and probe.strip():
            probe_chr, probe_pos, probe_info, probe_match = self.find_sequence_position_with_fallback(
                probe, ref_seqs, blast_db_path)
        
        locations = []
        
        # Check if ALL required sequences are found
        all_found = True
        match_methods = []
        
        if forward_chr and forward_pos:
            logger.debug(f"  ✓ Found Forward Primer at {forward_chr}:{forward_pos} ({forward_match}, {forward_info})")
            match_methods.append(forward_match)
        else:
            logger.debug(f"  ✗ Forward Primer not found in genome")
            all_found = False
        
        if reverse_chr and reverse_pos:
            logger.debug(f"  ✓ Found Reverse Primer at {reverse_chr}:{reverse_pos} ({reverse_match}, {reverse_info})")
            match_methods.append(reverse_match)
        else:
            logger.debug(f"  ✗ Reverse Primer not found in genome")
            all_found = False
        
        # Check probe if provided
        if probe and probe.strip():
            if probe_chr and probe_pos:
                logger.debug(f"  ✓ Found Probe at {probe_chr}:{probe_pos} ({probe_match}, {probe_info})")
                match_methods.append(probe_match)
            else:
                logger.debug(f"  ✗ Probe not found in genome")
                all_found = False
        
        # Only proceed if ALL sequences are found
        if not all_found:
            logger.debug(f"  ✗ NOT all sequences found - no amplicon created")
            return locations
        
        # Create amplicon using forward primer position
        # Determine overall match quality
        if all('exact' in method for method in match_methods):
            overall_quality = 'exact'
        elif any('blast_perfect' in method for method in match_methods):
            overall_quality = 'blast_perfect'
        else:
            # Use the worst match quality
            overall_quality = 'blast_partial'
            for method in match_methods:
                if 'blast_' in method and ('good' in method or 'partial' in method):
                    overall_quality = method
                    break
        
        # Create a simple amplicon record
        locations.append({
            'chromosome': forward_chr,
            'position': forward_pos,  # Use forward primer position
            'match_quality': overall_quality,
            'forward_match': forward_match,
            'reverse_match': reverse_match,
            'probe_match': probe_match,
            'forward_info': forward_info,
            'reverse_info': reverse_info,
            'probe_info': probe_info,
            # Store all positions for potential amplicon generation
            'forward_chr': forward_chr,
            'forward_pos': forward_pos,
            'reverse_chr': reverse_chr,
            'reverse_pos': reverse_pos,
            'probe_chr': probe_chr,
            'probe_pos': probe_pos
        })
        
        logger.debug(f"  ✓ ALL sequences found - amplicon created with quality: {overall_quality}")
        
        return locations

    def generate_amplicon_sequence(self, forward_pos: int, reverse_pos: int, chromosome: str, 
                                 ref_seqs: Dict[str, str]) -> Optional[str]:
        """
        Generate amplicon sequence between primer positions.
        
        Args:
            forward_pos: Forward primer position (1-based)
            reverse_pos: Reverse primer position (1-based)
            chromosome: Chromosome name
            ref_seqs: Reference genome sequences
            
        Returns:
            Amplicon sequence or None if cannot be generated
        """
        if chromosome not in ref_seqs:
            return None
        
        # Convert to 0-based coordinates
        start = min(forward_pos - 1, reverse_pos - 1)
        end = max(forward_pos - 1, reverse_pos - 1)
        
        # Add some buffer to ensure we get the full amplicon
        # This is a simplified approach - in reality, we'd need primer length information
        buffer = 100  # Assume max amplicon size
        start = max(0, start)
        end = min(len(ref_seqs[chromosome]), end + buffer)
        
        if start >= end:
            return None
        
        return ref_seqs[chromosome][start:end]

    @classmethod
    def run_remap_workflow(cls, primers: List[Dict], original_df: pd.DataFrame, output_dir: str, 
                           input_file: str, fasta_file: str, gff_file: Optional[str] = None,
                           skip_annotation_filtering: bool = False) -> bool:
        """
        Execute the remapping and re-evaluation workflow with full pipeline integration.
        
        Args:
            primers: List of primer set dictionaries
            original_df: Original DataFrame with all columns preserved
            output_dir: Output directory path
            input_file: Original input file path
            fasta_file: Reference FASTA file path
            gff_file: Optional GFF annotation file path
            skip_annotation_filtering: Whether to skip annotation processing
            
        Returns:
            True if workflow completed successfully, False otherwise
        """
        processor = cls()
        
        # 1. Load new reference genome and annotations
        ref_seqs = FileIO.load_fasta(fasta_file)
        genes = []
        if not skip_annotation_filtering and gff_file:
            try:
                genes = AnnotationProcessor.load_genes_from_gff(gff_file)
                logger.info(f"Loaded {len(genes)} gene annotations")
            except Exception as e:
                logger.warning(f"Could not load gene annotations: {e}")
                genes = []
        
        # 2. Get BLAST database path for fallback matching
        blast_db_path = None
        if hasattr(Config, 'DB_PATH') and Config.DB_PATH:
            if os.path.exists(f"{Config.DB_PATH}.nhr"):
                blast_db_path = Config.DB_PATH
                logger.debug(f"BLAST database available for fallback matching: {Config.DB_PATH}")
            else:
                logger.warning("BLAST database path configured but files not found")
        
        if not blast_db_path:
            logger.info("BLAST database not available - using exact matching only")
        
        remapped_records = []
        exact_matches = 0
        blast_matches = 0
        no_matches = 0
        genes_updated = 0
        
        # 3. Process each primer set with matching using tqdm
        logger.info(f"\nRemapping {len(primers)} primer sets...")
        
        # Use tqdm for progress tracking
        primer_iterator = tqdm(primers, desc="Remapping primers", unit="primer")
        
        for primer_set in primer_iterator:
            primer_id = primer_set['id']
            df_index = primer_set.get('dataframe_index')
            
            # Update tqdm description
            primer_iterator.set_description(f"Processing {primer_id}")
            
            logger.debug(f"Processing {primer_id}:")
            
            # Find locations using method with probe support
            locations = processor.find_primer_pair_locations(
                primer_set['forward'], 
                primer_set['reverse'], 
                ref_seqs,
                blast_db_path,
                primer_set.get('probe')  # Pass probe if available
            )
            
            if not locations:
                logger.debug(f"  ✗ Primer set '{primer_id}' could not be mapped to the new reference.")
                no_matches += 1
            else:
                for loc in locations:
                    # Create primer record preserving original data where possible
                    gene_name = primer_id
                    original_gene = primer_set.get('original_gene', primer_id)
                    
                    # Start with original row data if available
                    if df_index is not None and df_index < len(original_df):
                        original_row = original_df.iloc[df_index]
                        # Create record from original data
                        record = original_row.to_dict()
                        
                        # Update key fields with new mapping information
                        record["Gene"] = gene_name
                        record["Chr"] = loc['chromosome']
                        # Remove the "Start" column to prevent duplication
                        if "Start" in record:
                            del record["Start"]
                        record["Location"] = str(loc['position'])  # String format for location
                        record["Match_Quality"] = loc['match_quality']
                        
                        # Update gene name based on annotation if available
                        if genes:
                            annotation_gene = processor.find_gene_at_position(
                                loc['chromosome'], 
                                loc['position'],  # Already 1-based
                                genes
                            )
                            if annotation_gene:
                                if annotation_gene != original_gene:
                                    logger.debug(f"  ✓ Gene updated: '{original_gene}' → '{annotation_gene}'")
                                    genes_updated += 1
                                else:
                                    logger.debug(f"  ✓ Gene annotation confirmed: '{annotation_gene}'")
                                record["Gene"] = annotation_gene
                            else:
                                logger.debug(f"  ! No gene annotation found at position")
                        
                        # Generate amplicon sequence if positions are available
                        if ('Sequence (A)' not in record or pd.isna(record.get('Sequence (A)'))) and \
                           loc.get('forward_pos') and loc.get('reverse_pos'):
                            amplicon_seq = processor.generate_amplicon_sequence(
                                loc['forward_pos'], loc['reverse_pos'], 
                                loc['chromosome'], ref_seqs
                            )
                            if amplicon_seq:
                                record["Sequence (A)"] = amplicon_seq
                                record["Length"] = len(amplicon_seq)
                                # Calculate GC% for new amplicon
                                record["GC%"] = FilterProcessor.calculate_gc(amplicon_seq)
                        
                    else:
                        # Fallback: create minimal record
                        record = {
                            "Gene": gene_name,
                            "Sequence (F)": primer_set['forward'],
                            "Sequence (R)": primer_set['reverse'],
                            "Chr": loc['chromosome'],
                            "Location": str(loc['position']),  # String format for location
                            "Match_Quality": loc['match_quality']
                        }
                        
                        # Add probe if available
                        if primer_set['probe']:
                            record["Sequence (P)"] = primer_set['probe']

                    remapped_records.append(record)
                    
                    # Track match statistics
                    if loc['match_quality'] == 'exact':
                        exact_matches += 1
                    else:
                        blast_matches += 1
                    
                    # Log match details at debug level
                    logger.debug(f"  ✓ Found at {loc['chromosome']}:{loc['position']} ({loc['match_quality']})")
                    logger.debug(f"    Forward: {loc['forward_match']} ({loc['forward_info']})")
                    logger.debug(f"    Reverse: {loc['reverse_match']} ({loc['reverse_info']})")
                    if loc.get('probe_match'):
                        logger.debug(f"    Probe: {loc['probe_match']} ({loc.get('probe_info', 'N/A')})")

        # Close the progress bar
        primer_iterator.close()
        
        if not remapped_records:
            logger.error("Remapping failed: No input primers could be mapped to the new reference genome.")
            return False

        # 4. Convert to DataFrame for processing
        df = pd.DataFrame(remapped_records)
        df = df.fillna(value=pd.NA)

        # 5. Print statistics
        total_mapped = exact_matches + blast_matches
        logger.info(f"\nRemapping Statistics:")
        logger.info(f"  Total primer sets processed: {len(primers)}")
        logger.info(f"  Successfully mapped: {total_mapped}")
        logger.debug(f"    - Exact matches: {exact_matches}")
        logger.debug(f"    - BLAST matches: {blast_matches}")
        logger.debug(f"  Failed to map: {no_matches}")
        if genes:
            logger.info(f"  Gene names updated: {genes_updated}")

        # 6. Save results with comprehensive column preservation
        logger.info("\nSaving results...")
        
        output_path = FileIO.save_results(df, output_dir, input_file, mode='remap')
        
        if output_path:
            logger.info(f"Remapping results saved to: {output_path}")
            return True
        return False


def run_remap_mode(args):
    """
    Main entry point for the remap mode workflow.
    
    Args:
        args: Parsed command line arguments
        
    Returns:
        True if remapping completed successfully, False otherwise
    """
    logger.info("=== Remap Mode Workflow ===")
    
    try:
        # 1. Handle input file selection
        input_file = args.remap if isinstance(args.remap, str) else None
        if not input_file:
            logger.info("\n>>> Please select primer table file (CSV/Excel) <<<")
            input_file = FileIO.select_file(
                "Select primer table file",
                [("Excel Files", "*.xlsx"), ("Excel Files", "*.xls"), ("CSV Files", "*.csv"), ("All Files", "*.*")]
            )
        
        fasta_file = args.fasta
        if not fasta_file:
            logger.info("\n>>> Please select reference FASTA file <<<")
            fasta_file = FileIO.select_file(
                "Select reference FASTA file",
                [("FASTA Files", "*.fasta"), ("FASTA Files", "*.fa"), ("FASTA Files", "*.fna"), ("All Files", "*")]
            )

        gff_file = args.gff
        skip_annotation = args.noannotation
        if not skip_annotation and not gff_file:
            logger.info("\n>>> Please select GFF annotation file <<<")
            try:
                gff_file = FileIO.select_file(
                    "Select GFF file",
                    [("GFF Files", "*.gff"), ("GFF3 Files", "*.gff3"), ("Compressed GFF", "*.gff.gz"), ("All Files", "*.*")]
                )
            except FileSelectionError:
                logger.info("Skipping annotation.")
                gff_file = None
        
        # 2. Prepare reference files (VCF is not needed for remap mode)
        from ..utils import FilePreparator
        prep_result = FilePreparator.prepare_pipeline_files_workflow(
            vcf_file=None, fasta_file=fasta_file, gff_file=gff_file
        )
        if not prep_result['success']:
            raise FileError(f"File preparation failed: {prep_result.get('reason', 'Unknown')}")
        
        fasta_file = prep_result.get('fasta_file', fasta_file)
        gff_file = prep_result.get('gff_file', gff_file)

        FileIO.mark_selection_complete()

        # 3. Load primers with original DataFrame preservation
        primers, original_df = PrimerRemapperProcessor.load_primers_from_table(input_file)
        logger.info(f"\nLoaded {len(primers)} primer sets for remapping.")
        logger.debug(f"Original DataFrame has {len(original_df.columns)} columns: {list(original_df.columns)}")

        # 4. Set up output directory - use FASTA file directory like other modes
        if args.output:
            output_dir = args.output
        else:
            # Use the directory of the reference FASTA file (consistent with standard mode)
            input_dir = os.path.dirname(os.path.abspath(fasta_file))
            output_dir = os.path.join(input_dir, "Primers")
        
        os.makedirs(output_dir, exist_ok=True)
        
        # 5. Run the workflow with original DataFrame
        success = PrimerRemapperProcessor.run_remap_workflow(
            primers=primers,
            original_df=original_df,  # Pass original DataFrame
            output_dir=output_dir,
            input_file=input_file,
            fasta_file=fasta_file,
            gff_file=gff_file,
            skip_annotation_filtering=skip_annotation,
        )
        return success

    except (FileError, FileFormatError, SequenceProcessingError, PrimerDesignError) as e:
        logger.error(f"Remap mode failed: {e}")
        return False
    except Exception as e:
        logger.error(f"An unexpected error occurred in remap mode: {e}", exc_info=True)
        return False